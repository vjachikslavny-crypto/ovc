from __future__ import annotations

try:
    import audioop
except ModuleNotFoundError:
    audioop = None  # Removed in Python 3.13+
import csv
import gzip
import hashlib
import json
import os
import mimetypes
import logging
import subprocess
import tempfile
import itertools
import wave
import html as html_module
import zipfile
import xml.etree.ElementTree as ET
import shutil
from dataclasses import dataclass
from io import BytesIO
import io
from pathlib import Path
from typing import Optional, Tuple

from fastapi import HTTPException, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from PIL import Image, ImageDraw, ImageFont, ImageOps
from openpyxl import load_workbook
import xlrd
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.agent.block_models import (
    AudioBlock,
    AudioData,
    DocBlock,
    DocData,
    DocMeta,
    VideoBlock,
    VideoData,
    TableBlock,
    TableData,
    CodeBlock,
    CodeData,
    MarkdownBlock,
    MarkdownData,
    SlidesBlock,
    SlidesData,
    ImageBlock,
    ImageData,
    dump_block,
)
from app.db.models import FileAsset, generate_uuid

try:
    import bleach
except ImportError:  # pragma: no cover
    bleach = None

try:
    import mammoth
except ImportError:  # pragma: no cover
    mammoth = None

try:
    from striprtf.striprtf import rtf_to_text
except ImportError:  # pragma: no cover
    rtf_to_text = None

try:
    from mutagen import File as MutagenFile
except ImportError:  # pragma: no cover
    MutagenFile = None

try:
    from pypdf import PdfReader
except ImportError:  # pragma: no cover
    PdfReader = None

try:
    import fitz  # PyMuPDF
    HAS_PYMUPDF = True
except ImportError:  # pragma: no cover
    HAS_PYMUPDF = False
except Exception:
    HAS_PYMUPDF = False

try:
    from pdf2image import convert_from_bytes
    HAS_PDF2IMAGE = True
except ImportError:  # pragma: no cover
    HAS_PDF2IMAGE = False

# OVC: docx - Playwright больше не используется для генерации превью Word файлов
# Превью для Word файлов не генерируется, используется только inline просмотр


_resolved = Path(__file__).resolve()
# src/app/services/files.py -> parents[3] == project root (.../OVC)
PROJECT_ROOT = _resolved.parents[3] if len(_resolved.parents) >= 4 else Path.cwd()
_legacy_upload_root = Path.home() / "data" / "uploads"
_configured_upload_root = os.getenv("OVC_UPLOAD_ROOT", "").strip()

if _configured_upload_root:
    UPLOAD_ROOT = Path(_configured_upload_root).expanduser().resolve()
elif _legacy_upload_root.exists():
    # Keep backward compatibility with previously saved files.
    UPLOAD_ROOT = _legacy_upload_root
else:
    UPLOAD_ROOT = PROJECT_ROOT / "data" / "uploads"

ORIGINAL_DIR = UPLOAD_ROOT / "original"
PREVIEW_DIR = UPLOAD_ROOT / "preview"
PAGES_DIR = UPLOAD_ROOT / "pages"  # OVC: pdf - кэш страниц PDF
DOC_HTML_DIR = UPLOAD_ROOT / "doc_html"
WAVEFORM_DIR = UPLOAD_ROOT / "waveform"
SLIDES_DIR = UPLOAD_ROOT / "slides"
SLIDES_META_DIR = UPLOAD_ROOT / "slides_meta"
EXCEL_SUMMARY_DIR = UPLOAD_ROOT / "excel_summary"
EXCEL_CHARTS_DIR = UPLOAD_ROOT / "excel_charts"  # OVC: excel - диаграммы
EXCEL_CHARTS_META_DIR = UPLOAD_ROOT / "excel_charts_meta"  # OVC: excel - метаданные диаграмм
VIDEO_DIR = UPLOAD_ROOT / "videos"
CODE_DIR = UPLOAD_ROOT / "code"
MARKDOWN_DIR = UPLOAD_ROOT / "markdown"

for directory in (
    ORIGINAL_DIR,
    PREVIEW_DIR,
    PAGES_DIR,
    DOC_HTML_DIR,
    WAVEFORM_DIR,
    SLIDES_DIR,
    SLIDES_META_DIR,
    EXCEL_SUMMARY_DIR,
    EXCEL_CHARTS_DIR,
    EXCEL_CHARTS_META_DIR,
    VIDEO_DIR,
    CODE_DIR,
    MARKDOWN_DIR,
):
    directory.mkdir(parents=True, exist_ok=True)


IMAGE_MIME_TYPES = {
    "image/jpeg",
    "image/png",
    "image/webp",
    "image/gif",
}
PDF_MIME_TYPES = {"application/pdf"}
DOCX_MIME_TYPES = {
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
}
RTF_MIME_TYPES = {"application/rtf", "text/rtf"}
PPTX_MIME_TYPES = {"application/vnd.openxmlformats-officedocument.presentationml.presentation"}
EXCEL_MIME_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
    "text/csv",
}
EXCEL_EXTENSIONS = (".xlsx", ".xls", ".csv")
AUDIO_MIME_TYPES = {
    "audio/webm",
    "audio/webm;codecs=opus",
    "audio/ogg",
    "audio/ogg;codecs=opus",
    "audio/mpeg",
    "audio/wav",
    "audio/x-wav",
    "audio/wave",
    "audio/x-m4a",
    "audio/mp4",
    "audio/aac",
}
AUDIO_EXTENSIONS = (".webm", ".ogg", ".mp3", ".wav", ".m4a", ".aac")
VIDEO_MIME_TYPES = {
    "video/mp4",
    "video/webm",
    "video/quicktime",
    "video/x-matroska",
    "video/x-msvideo",
}
VIDEO_EXTENSIONS = (".mp4", ".webm", ".mov", ".mkv", ".avi")
MARKDOWN_EXTENSIONS = {".md", ".markdown"}
MARKDOWN_MIME_TYPES = {"text/markdown", "text/x-markdown", "text/plain"}
CODE_EXTENSIONS = {
    ".py",
    ".js",
    ".ts",
    ".tsx",
    ".jsx",
    ".json",
    ".html",
    ".htm",
    ".css",
    ".scss",
    ".yml",
    ".yaml",
    ".sh",
    ".bash",
    ".zsh",
    ".sql",
    ".c",
    ".h",
    ".cpp",
    ".hpp",
    ".java",
    ".kt",
    ".go",
    ".rs",
    ".php",
    ".rb",
    ".lua",
    ".r",
    ".tex",
    ".toml",
    ".ini",
    ".cfg",
    ".pl",
    ".cs",
    ".swift",
}

IMAGE_MAX_BYTES = 15 * 1024 * 1024
PDF_MAX_BYTES = 50 * 1024 * 1024
DOC_MAX_BYTES = 30 * 1024 * 1024
AUDIO_MAX_BYTES = 50 * 1024 * 1024
AUDIO_WAVE_POINTS = 256
SLIDES_MAX_BYTES = 50 * 1024 * 1024
SLIDES_TARGET_WIDTH = 1600
EXCEL_MAX_BYTES = 40 * 1024 * 1024
EXCEL_PREVIEW_ROWS = 5
EXCEL_WINDOW_LIMIT = 1000
EXCEL_CHARTS_TARGET_WIDTH = 1600
VIDEO_MAX_BYTES = 200 * 1024 * 1024
MAX_CODE_LINES = 10_000
CODE_PREVIEW_LINES = 300
CODE_MAX_BYTES = 5 * 1024 * 1024
MARKDOWN_PREVIEW_MAX_BYTES = 200_000
CODE_GZIP_THRESHOLD = 10 * 1024 * 1024
_FFMPEG_PATH = shutil.which("ffmpeg")
_FFPROBE_PATH = shutil.which("ffprobe")


def ffmpeg_available() -> bool:
    return _FFMPEG_PATH is not None


def _ffprobe_available() -> bool:
    return _FFPROBE_PATH is not None


CODE_LANGUAGE_MAP = {
    ".py": "python",
    ".js": "javascript",
    ".ts": "typescript",
    ".tsx": "tsx",
    ".jsx": "jsx",
    ".json": "json",
    ".md": "markdown",
    ".markdown": "markdown",
    ".html": "markup",
    ".htm": "markup",
    ".css": "css",
    ".scss": "css",
    ".yml": "yaml",
    ".yaml": "yaml",
    ".sh": "bash",
    ".bash": "bash",
    ".zsh": "bash",
    ".sql": "sql",
    ".c": "c",
    ".h": "c",
    ".cpp": "cpp",
    ".hpp": "cpp",
    ".java": "java",
    ".kt": "kotlin",
    ".go": "go",
    ".rs": "rust",
    ".php": "php",
    ".rb": "ruby",
    ".lua": "lua",
    ".r": "r",
    ".tex": "latex",
    ".toml": "toml",
    ".ini": "ini",
    ".cfg": "ini",
    ".pl": "perl",
    ".cs": "csharp",
    ".swift": "swift",
}


def _detect_code_language(extension: str) -> str:
    return CODE_LANGUAGE_MAP.get(extension.lower(), "plaintext")


def count_file_lines(path: Path) -> int:
    count = 0
    with path.open("r", encoding="utf-8", errors="ignore") as handle:
        for _ in handle:
            count += 1
    return count

ALLOWED_HTML_TAGS = [
    "p",
    "br",
    "strong",
    "em",
    "u",
    "s",
    "a",
    "ul",
    "ol",
    "li",
    "blockquote",
    "code",
    "pre",
    "h1",
    "h2",
    "h3",
    "h4",
    "h5",
    "h6",
    "table",
    "thead",
    "tbody",
    "tr",
    "td",
    "th",
    "img",
    "figure",
    "figcaption",
    "span",
]

ALLOWED_HTML_ATTRS = {
    "a": ["href", "title", "rel", "target"],
    "img": ["src", "alt"],
    "td": ["colspan", "rowspan"],
    "th": ["colspan", "rowspan"],
    "*": ["class"],
}


class FileMetadata(BaseModel):
    kind: str
    mime: str
    extension: str
    max_bytes: int


def _guess_extension(filename: str, mime: str) -> str:
    suffix = Path(filename or "").suffix.lower()
    if suffix:
        return suffix
    guessed = mimetypes.guess_extension(mime)
    return guessed or ".bin"


def _audio_mime_from_filename(filename: str) -> str:
    suffix = Path(filename or "").suffix.lower()
    if suffix == ".mp3":
        return "audio/mpeg"
    if suffix == ".wav":
        return "audio/wav"
    if suffix == ".m4a":
        return "audio/mp4"
    if suffix == ".aac":
        return "audio/aac"
    if suffix == ".ogg":
        return "audio/ogg"
    return "audio/webm"


def _classify_file(upload: UploadFile) -> FileMetadata:
    filename = upload.filename or ""
    mime = (upload.content_type or "").lower()
    if not mime:
        guessed, _ = mimetypes.guess_type(filename)
        mime = (guessed or "application/octet-stream").lower()

    if mime in IMAGE_MIME_TYPES or filename.lower().endswith((".jpg", ".jpeg", ".png", ".webp", ".gif")):
        return FileMetadata(kind="image", mime=mime, extension=_guess_extension(filename, mime), max_bytes=IMAGE_MAX_BYTES)
    if mime in PDF_MIME_TYPES or filename.lower().endswith(".pdf"):
        return FileMetadata(kind="pdf", mime="application/pdf", extension=".pdf", max_bytes=PDF_MAX_BYTES)
    if mime in DOCX_MIME_TYPES or filename.lower().endswith(".docx"):
        return FileMetadata(kind="docx", mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document", extension=".docx", max_bytes=DOC_MAX_BYTES)
    if mime in RTF_MIME_TYPES or filename.lower().endswith(".rtf"):
        return FileMetadata(kind="rtf", mime="application/rtf", extension=".rtf", max_bytes=DOC_MAX_BYTES)
    if mime in PPTX_MIME_TYPES or filename.lower().endswith(".pptx"):
        return FileMetadata(kind="pptx", mime="application/vnd.openxmlformats-officedocument.presentationml.presentation", extension=".pptx", max_bytes=SLIDES_MAX_BYTES)
    if mime in EXCEL_MIME_TYPES or filename.lower().endswith(EXCEL_EXTENSIONS):
        lower = filename.lower()
        if lower.endswith(".csv") or mime == "text/csv":
            excel_kind = "csv"
            extension = ".csv"
            excel_mime = "text/csv"
        elif lower.endswith(".xls"):
            excel_kind = "xls"
            extension = ".xls"
            excel_mime = "application/vnd.ms-excel"
        else:
            excel_kind = "xlsx"
            extension = ".xlsx"
            excel_mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return FileMetadata(kind=excel_kind, mime=excel_mime, extension=extension, max_bytes=EXCEL_MAX_BYTES)
    if mime in AUDIO_MIME_TYPES or filename.lower().endswith(AUDIO_EXTENSIONS):
        audio_mime = mime if mime in AUDIO_MIME_TYPES else _audio_mime_from_filename(filename)
        extension = _guess_extension(filename, audio_mime) or ".webm"
        return FileMetadata(kind="audio", mime=audio_mime, extension=extension, max_bytes=AUDIO_MAX_BYTES)
    if mime in VIDEO_MIME_TYPES or filename.lower().endswith(VIDEO_EXTENSIONS):
        video_mime = mime if mime in VIDEO_MIME_TYPES else "video/mp4"
        extension = _guess_extension(filename, video_mime) or ".mp4"
        return FileMetadata(kind="video", mime=video_mime, extension=extension, max_bytes=VIDEO_MAX_BYTES)
    suffix = Path(filename or "").suffix.lower()
    if suffix in CODE_EXTENSIONS:
        code_mime = mime if mime.startswith("text/") else "text/plain"
        extension = suffix or ".txt"
        return FileMetadata(kind="code", mime=code_mime, extension=extension, max_bytes=CODE_MAX_BYTES)
    if suffix in MARKDOWN_EXTENSIONS or mime in MARKDOWN_MIME_TYPES:
        md_mime = "text/markdown"
        extension = suffix or ".md"
        return FileMetadata(kind="markdown", mime=md_mime, extension=extension, max_bytes=CODE_MAX_BYTES)

    raise HTTPException(
        status_code=415,
        detail=(
            "Unsupported file type. Supported: png/jpg/jpeg/gif/webp, pdf, "
            "docx, xlsx, pptx, mp3/wav/m4a, mp4/mov, plus markdown/code formats."
        ),
    )


def _write_file(path: Path, data: bytes) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("wb") as handle:
        handle.write(data)


def _slides_dir_path(file_id: str) -> Path:
    return SLIDES_DIR / file_id


def _slides_json_path(file_id: str) -> Path:
    return SLIDES_META_DIR / f"{file_id}.json"


def _excel_charts_dir_path(file_id: str) -> Path:
    return EXCEL_CHARTS_DIR / file_id


def _excel_charts_json_path(file_id: str) -> Path:
    return EXCEL_CHARTS_META_DIR / f"{file_id}.json"


VIDEO_POSTER_NAME = "poster.webp"


def _video_dir(file_id: str) -> Path:
    return VIDEO_DIR / file_id


def _video_original_path(file_id: str, extension: str) -> Path:
    return _video_dir(file_id) / f"original{extension}"


def _video_poster_path(file_id: str) -> Path:
    return _video_dir(file_id) / VIDEO_POSTER_NAME


def _video_source_url(file_id: str) -> str:
    return f"/files/{file_id}/video/source"


def _video_poster_url(file_id: str) -> str:
    return f"/files/{file_id}/video/poster.webp"


def _code_dir(file_id: str) -> Path:
    return CODE_DIR / file_id


def _code_original_path(file_id: str, extension: str) -> Path:
    return _code_dir(file_id) / f"original{extension}"


def _code_preview_url(file_id: str, max_lines: int = CODE_PREVIEW_LINES) -> str:
    return f"/files/{file_id}/code/preview?maxLines={max_lines}"


def _code_raw_url(file_id: str) -> str:
    return f"/files/{file_id}/code/raw"


def _code_meta_url(file_id: str) -> str:
    return f"/files/{file_id}/code/meta"


def _markdown_dir(file_id: str) -> Path:
    return MARKDOWN_DIR / file_id


def _markdown_raw_path(file_id: str, extension: str) -> Path:
    return _markdown_dir(file_id) / f"original{extension}"


def _markdown_raw_url(file_id: str) -> str:
    return f"/files/{file_id}/md/raw"


def _markdown_preview_url(file_id: str, max_bytes: int = MARKDOWN_PREVIEW_MAX_BYTES) -> str:
    return f"/files/{file_id}/md/preview?maxBytes={max_bytes}"


def _generate_image_preview(data: bytes) -> Tuple[bytes, int, int]:
    buffer = BytesIO(data)
    with Image.open(buffer) as img:
        img = ImageOps.exif_transpose(img)
        width, height = img.size
        preview = img.copy()
        preview.thumbnail((1600, 1600))
        out = BytesIO()
        preview.save(out, format="WEBP", quality=85)
        preview_bytes = out.getvalue()
    return preview_bytes, width, height


def extract_video_meta(path: Path) -> dict:
    meta = {"duration": None, "width": None, "height": None}
    if not _ffprobe_available():
        return meta
    try:
        result = subprocess.run(
            [
                _FFPROBE_PATH,
                "-v",
                "error",
                "-select_streams",
                "v:0",
                "-show_entries",
                "stream=width,height,duration",
                "-of",
                "json",
                str(path),
            ],
            capture_output=True,
            text=True,
            check=True,
        )
        payload = json.loads(result.stdout or "{}")
        stream = (payload.get("streams") or [{}])[0]
        duration = stream.get("duration")
        width = stream.get("width")
        height = stream.get("height")
        if duration is not None:
            try:
                meta["duration"] = float(duration)
            except (ValueError, TypeError):
                meta["duration"] = None
        if width is not None:
            try:
                meta["width"] = int(width)
            except (ValueError, TypeError):
                meta["width"] = None
        if height is not None:
            try:
                meta["height"] = int(height)
            except (ValueError, TypeError):
                meta["height"] = None
        return meta
    except Exception as exc:
        logger = logging.getLogger(__name__)
        logger.warning(f"ffprobe failed for video {path.name}: {exc}")
        return meta


def render_video_poster(video_path: Path, out_path: Path, *, at_percent: float = 0.2, duration: Optional[float] = None) -> bool:
    if not ffmpeg_available():
        return False
    seek = duration * at_percent if duration and duration > 0 else 1.0
    try:
        subprocess.run(
            [
                _FFMPEG_PATH,
                "-y",
                "-v",
                "error",
                "-ss",
                str(max(seek, 0.01)),
                "-i",
                str(video_path),
                "-frames:v",
                "1",
                "-vf",
                "scale='min(1280,iw)':-2",
                str(out_path),
            ],
            check=True,
        )
        return True
    except Exception as exc:
        logger = logging.getLogger(__name__)
        logger.warning(f"ffmpeg poster generation failed for {video_path.name}: {exc}")
        return False


def _render_pdf_page(data: bytes, page_num: int, scale: float = 1.0) -> Optional[bytes]:
    """OVC: pdf - рендеринг страницы PDF в изображение."""
    if page_num < 1:
        return None
    
    try:
        # Попытка использовать PyMuPDF (fitz) - самый быстрый вариант
        if HAS_PYMUPDF:
            try:
                doc = fitz.open(stream=data, filetype="pdf")
                if page_num > len(doc):
                    doc.close()
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.warning(f"Page {page_num} out of range (PDF has {len(doc)} pages)")
                    return None
                page = doc[page_num - 1]  # 0-based index
                # PyMuPDF: zoom factor (1.0 = 72 DPI, 2.0 = 144 DPI)
                # Для лучшего качества используем scale * 1.5
                zoom = scale * 1.5
                mat = fitz.Matrix(zoom, zoom)
                pix = page.get_pixmap(matrix=mat, alpha=False)
                # Конвертируем в PNG сначала
                img_data = pix.tobytes("png")
                doc.close()
                # Конвертируем PNG в WEBP
                img = Image.open(BytesIO(img_data))
                out = BytesIO()
                img.save(out, format="WEBP", quality=85, method=6)
                return out.getvalue()
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"PyMuPDF render error for page {page_num}: {e}", exc_info=True)
                return None
        
        # Попытка использовать pdf2image
        if HAS_PDF2IMAGE:
            try:
                images = convert_from_bytes(data, first_page=page_num, last_page=page_num, dpi=int(72 * scale))
                if not images:
                    return None
                img = images[0]
                # Масштабируем если нужно
                if scale != 1.0:
                    new_size = (int(img.width * scale), int(img.height * scale))
                    img = img.resize(new_size, Image.Resampling.LANCZOS)
                out = BytesIO()
                img.save(out, format="WEBP", quality=85)
                return out.getvalue()
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"pdf2image render error: {e}", exc_info=True)
                return None
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"PDF render error: {e}", exc_info=True)
        return None
    
    return None


def _generate_pdf_preview(data: bytes, file_id: str, pages: Optional[int]) -> bytes:
    """OVC: pdf - генерация preview для PDF (рендер первой страницы или placeholder)."""
    # Пытаемся отрендерить первую страницу
    try:
        preview_img = _render_pdf_page(data, 1, scale=1.0)
        if preview_img:
            # Масштабируем до разумного размера для preview
            img = Image.open(BytesIO(preview_img))
            img.thumbnail((1200, 1600), Image.Resampling.LANCZOS)
            out = BytesIO()
            img.save(out, format="WEBP", quality=85)
            return out.getvalue()
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"PDF preview generation error: {e}", exc_info=True)
    
    # Fallback: placeholder если рендеринг недоступен
    import logging
    logger = logging.getLogger(__name__)
    logger.warning(f"PDF preview: using placeholder for {file_id} (libraries: PyMuPDF={HAS_PYMUPDF}, pdf2image={HAS_PDF2IMAGE})")
    width, height = 1200, 800
    image = Image.new("RGB", (width, height), "#f8fafc")
    draw = ImageDraw.Draw(image)
    font = ImageFont.load_default()
    title = "PDF Document"
    subtitle = f"{pages or '?'} pages" if pages else "PDF файл"
    # Простой расчет позиции для текста
    try:
        bbox = draw.textbbox((0, 0), title, font=font)
        title_width = bbox[2] - bbox[0]
        title_height = bbox[3] - bbox[1]
    except:
        title_width = len(title) * 20
        title_height = 30
    try:
        bbox_small = draw.textbbox((0, 0), subtitle, font=font)
        subtitle_width = bbox_small[2] - bbox_small[0]
    except:
        subtitle_width = len(subtitle) * 12
    draw.text((width // 2 - title_width // 2, height // 2 - 40), title, fill="#0f172a", font=font)
    draw.text((width // 2 - subtitle_width // 2, height // 2 + 10), subtitle, fill="#475569", font=font)
    out = BytesIO()
    image.save(out, format="WEBP", quality=85)
    return out.getvalue()


def _pdf_page_count(data: bytes) -> Optional[int]:
    if PdfReader is None:
        return None
    try:
        reader = PdfReader(BytesIO(data))
        return len(reader.pages)
    except Exception:
        return None


def _sanitize_html_fragment(fragment: str) -> str:
    if bleach is None:  # pragma: no cover
        return html_module.escape(fragment)
    cleaned = bleach.clean(fragment, tags=ALLOWED_HTML_TAGS, attributes=ALLOWED_HTML_ATTRS, strip=True)
    return cleaned


def _write_doc_html(file_id: str, html_fragment: str) -> Path:
    path = DOC_HTML_DIR / f"{file_id}.html"
    path.write_text(html_fragment, encoding="utf-8")
    return path


# OVC: docx - функция генерации превью удалена, т.к. превью для Word файлов не генерируется
# Вместо превью используется информативный бейдж с названием файла и количеством слов


def _convert_docx_to_html(data: bytes) -> tuple[str, Optional[int]]:
    if mammoth is None:
        raise HTTPException(status_code=500, detail="DOCX preview is unavailable (mammoth not installed).")
    result = mammoth.convert_to_html(BytesIO(data))
    html_fragment = result.value or ""
    if bleach is not None:
        text_only = bleach.clean(html_fragment, tags=[], strip=True)
    else:  # pragma: no cover
        text_only = html_module.escape(html_fragment)
    word_count = len(text_only.split())
    return html_fragment, word_count


def _convert_rtf_to_html(data: bytes) -> tuple[str, Optional[int]]:
    if rtf_to_text is None:
        raise HTTPException(status_code=500, detail="RTF preview is unavailable (striprtf not installed).")
    try:
        text = rtf_to_text(data.decode("utf-8", errors="ignore"))
    except Exception:
        text = rtf_to_text(data.decode("latin-1", errors="ignore"))
    paragraphs = [seg.strip() for seg in text.split("\n") if seg.strip()]
    html_parts = []
    for paragraph in paragraphs:
        html_parts.append(f"<p>{html_module.escape(paragraph)}</p>")
    html_fragment = "".join(html_parts)
    word_count = len(text.split())
    return html_fragment, word_count


def _default_waveform(points: int = 64) -> list[float]:
    return [0.2 for _ in range(points)]


def _extract_audio_metadata(data: bytes, mime: str) -> tuple[Optional[float], list[float]]:
    duration = None
    waveform: list[float] = []
    if MutagenFile is not None:
        try:
            audio = MutagenFile(BytesIO(data))
            if audio and audio.info:
                duration = getattr(audio.info, "length", None)
        except Exception:
            duration = None

    # Waveform only for PCM wav; others fallback
    try:
        if mime in {"audio/wav", "audio/x-wav", "audio/wave"}:
            with wave.open(BytesIO(data)) as wav_file:
                frames = wav_file.readframes(wav_file.getnframes())
                sample_width = wav_file.getsampwidth()
                channels = wav_file.getnchannels()
                if channels > 1:
                    if audioop is not None:
                        frames = audioop.tomono(frames, sample_width, 0.5, 0.5)
                    else:
                        # Pure-Python mono downmix: average channels
                        import struct
                        fmt = {1: 'b', 2: '<h', 4: '<i'}.get(sample_width, '<h')
                        frame_size = sample_width * channels
                        mono_frames = bytearray()
                        for i in range(0, len(frames), frame_size):
                            total = 0
                            for ch in range(channels):
                                offset = i + ch * sample_width
                                sample = struct.unpack_from(fmt, frames, offset)[0]
                                total += sample
                            avg = total // channels
                            mono_frames.extend(struct.pack(fmt, avg))
                        frames = bytes(mono_frames)
                total_samples = len(frames) // sample_width
                if total_samples == 0:
                    raise ValueError("empty audio")
                bucket = max(total_samples // AUDIO_WAVE_POINTS, 1)
                waveform = []
                max_sample = float((1 << (8 * sample_width - 1)) - 1) or 1.0
                for idx in range(AUDIO_WAVE_POINTS):
                    start = idx * bucket * sample_width
                    if start >= len(frames):
                        break
                    end = min(len(frames), start + bucket * sample_width)
                    chunk = frames[start:end]
                    if not chunk:
                        waveform.append(0.0)
                        continue
                    if audioop is not None:
                        peak = audioop.max(chunk, sample_width)
                    else:
                        # Pure-Python peak detection
                        import struct
                        fmt = {1: 'b', 2: '<h', 4: '<i'}.get(sample_width, '<h')
                        peak = 0
                        for si in range(0, len(chunk), sample_width):
                            val = abs(struct.unpack_from(fmt, chunk, si)[0])
                            if val > peak:
                                peak = val
                    waveform.append(round(min(1.0, peak / max_sample), 4))
    except Exception:
        waveform = []

    if not waveform:
        waveform = _default_waveform()
    return duration, waveform


def _write_waveform(file_id: str, values: list[float]) -> Path:
    path = WAVEFORM_DIR / f"{file_id}.json"
    path.write_text(json.dumps(values), encoding="utf-8")
    return path


def _render_slide_images(pdf_bytes: bytes, file_id: str, pages: int) -> tuple[Path, Path, int, Path]:
    logger = logging.getLogger(__name__)
    slides_dir = _slides_dir_path(file_id)
    slides_dir.mkdir(parents=True, exist_ok=True)
    preview_path: Optional[Path] = None
    width = height = None
    for page in range(1, pages + 1):
        slide_bytes = _render_pdf_page(pdf_bytes, page, scale=1.2)
        if not slide_bytes:
            raise HTTPException(status_code=500, detail="Failed to render PPTX slide")
        image = Image.open(BytesIO(slide_bytes))
        if SLIDES_TARGET_WIDTH and image.width > SLIDES_TARGET_WIDTH:
            image.thumbnail((SLIDES_TARGET_WIDTH, SLIDES_TARGET_WIDTH * 2), Image.Resampling.LANCZOS)
            buffer = BytesIO()
            image.save(buffer, format="WEBP", quality=85)
            slide_bytes = buffer.getvalue()
        slide_path = slides_dir / f"{page}.webp"
        _write_file(slide_path, slide_bytes)
        if preview_path is None:
            preview_path = slide_path
            width, height = image.size
    if preview_path is None:
        raise HTTPException(status_code=500, detail="No slides rendered from PPTX")
    meta = {
        "count": pages,
        "format": "webp",
        "width": width,
        "height": height,
    }
    json_path = _slides_json_path(file_id)
    json_path.write_text(json.dumps(meta), encoding="utf-8")
    logger.info("Generated %s slides for %s", pages, file_id)
    return preview_path, json_path, pages, slides_dir


def _find_libreoffice() -> Optional[str]:
    """Находит путь к LibreOffice soffice на разных системах."""
    # Стандартные пути для macOS
    mac_paths = [
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
        "/opt/homebrew/bin/soffice",
        "/usr/local/bin/soffice",
    ]
    # Проверяем стандартные пути
    for path in mac_paths:
        if Path(path).exists():
            return path
    # Проверяем PATH
    result = subprocess.run(["which", "soffice"], capture_output=True, text=True)
    if result.returncode == 0:
        return result.stdout.strip()
    return None


def _convert_pptx_to_slides(original_path: Path, file_id: str) -> tuple[Path, Path, int, Path]:
    logger = logging.getLogger(__name__)
    soffice_path = _find_libreoffice()
    if not soffice_path:
        raise HTTPException(
            status_code=500,
            detail="LibreOffice (soffice) is not installed. Please install LibreOffice from https://www.libreoffice.org/download/"
        )
    try:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            cmd = [
                soffice_path,
                "--headless",
                "--convert-to",
                "pdf",
                "--outdir",
                str(tmp_path),
                str(original_path),
            ]
            result = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, check=True)
            logger.info("LibreOffice output: %s", result.stdout.decode(errors="ignore").strip())
            pdf_path = tmp_path / f"{original_path.stem}.pdf"
            if not pdf_path.exists():
                raise HTTPException(status_code=500, detail="Failed to convert PPTX to PDF")
            pdf_bytes = pdf_path.read_bytes()
    except subprocess.CalledProcessError as exc:
        error_msg = exc.stderr.decode(errors='ignore') if exc.stderr else str(exc)
        raise HTTPException(status_code=500, detail=f"LibreOffice failed: {error_msg}")

    pages = _pdf_page_count(pdf_bytes)
    if not pages:
        raise HTTPException(status_code=500, detail="PPTX contains no slides")
    return _render_slide_images(pdf_bytes, file_id, pages)


def _clean_cell(value) -> str:
    if value is None:
        return ""
    return str(value)


def _normalize_header(header: Optional[tuple], fallback: int) -> list[str]:
    values = [_clean_cell(v) for v in (header or [])]
    length = max(len(values), fallback or 0)
    if length == 0:
        length = len(values)
    if length == 0:
        length = 1
    if not any(values):
        return [f"Column {i + 1}" for i in range(length)]
    while len(values) < length:
        values.append("")
    return values[:length]


def _format_row(row: tuple, columns: int) -> list[str]:
    values = [_clean_cell(v) for v in (row or [])]
    if len(values) < columns:
        values.extend([""] * (columns - len(values)))
    return values[:columns]


def _summarize_xlsx(path: Path) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    all_sheets = []
    for ws in wb.worksheets:
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            header = []
        columns = _normalize_header(header, ws.max_column)
        preview = []
        for row in itertools.islice(rows_iter, EXCEL_PREVIEW_ROWS):
            preview.append(_format_row(row, len(columns)))
        total_rows = max(ws.max_row - (1 if columns else 0), 0)
        all_sheets.append(
            {
                "name": ws.title,
                "rows": total_rows,
                "cols": len(columns),
                "columns": columns,
                "preview": preview,
            }
        )
    
    # Фильтруем листы: показываем все непустые + максимум 2 следующих пустых
    sheets = []
    empty_count = 0
    for sheet in all_sheets:
        if sheet["rows"] > 0:
            # Непустой лист - всегда показываем
            sheets.append(sheet)
            empty_count = 0  # Сбрасываем счетчик пустых
        else:
            # Пустой лист - показываем только если это один из первых двух пустых после непустых
            if empty_count < 2:
                sheets.append(sheet)
                empty_count += 1
            # Если уже показали 2 пустых, остальные пропускаем
    
    default_sheet = sheets[0]["name"] if sheets else None
    return {"sheets": sheets, "defaultSheet": default_sheet}


def _summarize_xls(path: Path) -> dict:
    book = xlrd.open_workbook(path, on_demand=True)
    all_sheets = []
    for sheet_name in book.sheet_names():
        sh = book.sheet_by_name(sheet_name)
        header = sh.row_values(0) if sh.nrows else []
        columns = _normalize_header(header, sh.ncols)
        preview = []
        start_row = 1 if sh.nrows > 0 else 0
        for row_idx in range(start_row, min(sh.nrows, start_row + EXCEL_PREVIEW_ROWS)):
            preview.append(_format_row(sh.row_values(row_idx), len(columns)))
        total_rows = max(sh.nrows - (1 if columns else 0), 0)
        all_sheets.append(
            {
                "name": sheet_name,
                "rows": total_rows,
                "cols": len(columns),
                "columns": columns,
                "preview": preview,
            }
        )
    
    # Фильтруем листы: показываем все непустые + максимум 2 следующих пустых
    sheets = []
    empty_count = 0
    for sheet in all_sheets:
        if sheet["rows"] > 0:
            # Непустой лист - всегда показываем
            sheets.append(sheet)
            empty_count = 0  # Сбрасываем счетчик пустых
        else:
            # Пустой лист - показываем только если это один из первых двух пустых после непустых
            if empty_count < 2:
                sheets.append(sheet)
                empty_count += 1
            # Если уже показали 2 пустых, остальные пропускаем
    
    default_sheet = sheets[0]["name"] if sheets else None
    return {"sheets": sheets, "defaultSheet": default_sheet}


def _summarize_csv(path: Path) -> dict:
    with path.open("r", newline="", encoding="utf-8", errors="ignore") as handle:
        reader = csv.reader(handle)
        header = next(reader, [])
        columns = _normalize_header(header, len(header))
        preview = []
        total_rows = 0
        for row in reader:
            total_rows += 1
            if len(preview) < EXCEL_PREVIEW_ROWS:
                preview.append(_format_row(row, len(columns)))
    sheet_name = path.stem or "CSV"
    sheet = {
        "name": sheet_name,
        "rows": total_rows,
        "cols": len(columns),
        "columns": columns,
        "preview": preview,
    }
    return {"sheets": [sheet], "defaultSheet": sheet_name}


def _prepare_excel_summary(kind: str, original_path: Path, file_id: str) -> tuple[Path, Optional[str]]:
    if kind == "csv":
        summary = _summarize_csv(original_path)
    elif kind == "xls":
        summary = _summarize_xls(original_path)
    else:
        summary = _summarize_xlsx(original_path)
    summary["maxWindow"] = EXCEL_WINDOW_LIMIT
    summary_path = _excel_summary_path(file_id)
    summary_path.write_text(json.dumps(summary, ensure_ascii=False), encoding="utf-8")
    return summary_path, summary.get("defaultSheet")


def _load_excel_summary(asset: FileAsset) -> dict:
    if not asset.path_excel_summary:
        raise HTTPException(status_code=404, detail="Excel summary not available")
    path = Path(asset.path_excel_summary)
    if not path.exists():
        raise HTTPException(status_code=202, detail="Summary is still processing")
    return json.loads(path.read_text(encoding="utf-8"))


def _read_excel_window(asset: FileAsset, sheet_name: str, offset: int, limit: int) -> dict:
    summary = _load_excel_summary(asset)
    sheet_meta = next((sheet for sheet in summary.get("sheets", []) if sheet["name"] == sheet_name), None)
    if not sheet_meta:
        raise HTTPException(status_code=404, detail="Sheet not found")
    total_rows = max(int(sheet_meta.get("rows") or 0), 0)
    columns = sheet_meta.get("columns") or []
    if offset < 0 or offset > total_rows:
        raise HTTPException(status_code=400, detail="Invalid offset")
    limit = max(1, min(limit, EXCEL_WINDOW_LIMIT))
    path = Path(asset.path_original)
    if not path.exists():
        raise HTTPException(status_code=404, detail="Original file missing")

    rows: list[list[str]] = []
    if asset.kind in {"xlsx"}:
        wb = load_workbook(path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            raise HTTPException(status_code=404, detail="Sheet not found")
        ws = wb[sheet_name]
        start_row = offset + 2  # skip header
        end_row = min(ws.max_row, start_row + limit - 1)
        if start_row > ws.max_row:
            start_row = ws.max_row + 1
        for row in ws.iter_rows(min_row=start_row, max_row=end_row, values_only=True):
            rows.append(_format_row(row, len(columns)))
    elif asset.kind == "xls":
        book = xlrd.open_workbook(path, on_demand=True)
        if sheet_name not in book.sheet_names():
            raise HTTPException(status_code=404, detail="Sheet not found")
        sh = book.sheet_by_name(sheet_name)
        start_row = offset + 1
        end_row = min(sh.nrows, start_row + limit)
        for row_idx in range(start_row, end_row):
            rows.append(_format_row(sh.row_values(row_idx), len(columns)))
    elif asset.kind == "csv":
        with path.open("r", newline="", encoding="utf-8", errors="ignore") as handle:
            reader = csv.reader(handle)
            next(reader, None)  # header
            for _ in range(offset):
                try:
                    next(reader)
                except StopIteration:
                    break
            for _ in range(limit):
                try:
                    row = next(reader)
                except StopIteration:
                    break
                rows.append(_format_row(row, len(columns)))
    else:
        raise HTTPException(status_code=400, detail="File is not a table")

    return {
        "offset": offset,
        "limit": limit,
        "total": total_rows,
        "columns": columns,
        "rows": rows,
        "sheet": sheet_name,
    }


def _iter_sheet_csv(asset: FileAsset, sheet_name: str):
    path = Path(asset.path_original)
    if asset.kind == "csv":
        return FileResponse(path, media_type="text/csv", filename=asset.filename)

    if asset.kind == "xlsx":
        wb = load_workbook(path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            raise HTTPException(status_code=404, detail="Sheet not found")
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        return _stream_rows_as_csv(rows_iter, sheet_name)

    if asset.kind == "xls":
        book = xlrd.open_workbook(path, on_demand=True)
        if sheet_name not in book.sheet_names():
            raise HTTPException(status_code=404, detail="Sheet not found")
        sh = book.sheet_by_name(sheet_name)

        def row_iter():
            for idx in range(sh.nrows):
                yield sh.row_values(idx)

        return _stream_rows_as_csv(row_iter(), sheet_name)

    raise HTTPException(status_code=400, detail="File is not a table")


def _stream_rows_as_csv(rows_iter, sheet_name: str) -> StreamingResponse:
    def generator():
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        for row in rows_iter:
            buffer.seek(0)
            buffer.truncate(0)
            writer.writerow([_clean_cell(v) for v in (row or [])])
            data = buffer.getvalue()
            yield data.encode("utf-8")

    headers = {"Content-Disposition": f'attachment; filename="{sheet_name}.csv"'}
    return StreamingResponse(generator(), media_type="text/csv", headers=headers)


def read_markdown_preview(asset: FileAsset, max_bytes: int) -> tuple[str, bool]:
    path = _get_markdown_file_path(asset)
    max_bytes = max(1, max_bytes)
    truncated = False
    total = 0
    chunks: list[str] = []
    with path.open("r", encoding="utf-8", errors="replace") as handle:
        for line in handle:
            encoded = line.encode("utf-8")
            if total + len(encoded) > max_bytes:
                truncated = True
                break
            chunks.append(line)
            total += len(encoded)
    return "".join(chunks), truncated


def read_code_segment(asset: FileAsset, start: int, max_lines: int) -> tuple[str, bool]:
    path = _get_code_file_path(asset)
    start = max(0, start)
    max_lines = max(1, min(max_lines, MAX_CODE_LINES))
    lines: list[str] = []
    truncated = False
    with path.open("r", encoding="utf-8", errors="replace") as handle:
        for _ in range(start):
            if handle.readline() == "":
                return "", False
        for _ in range(max_lines):
            chunk = handle.readline()
            if chunk == "":
                break
            lines.append(chunk)
        else:
            if handle.readline():
                truncated = True
    if not truncated and asset.code_line_count is not None:
        truncated = start + len(lines) < asset.code_line_count
    return "".join(lines), truncated


def prepare_code_bytes(text: str, asset: FileAsset) -> tuple[bytes, dict]:
    data = text.encode("utf-8")
    headers: dict[str, str] = {}
    size_hint = asset.size or len(data)
    if size_hint > CODE_GZIP_THRESHOLD:
        buffer = BytesIO()
        with gzip.GzipFile(fileobj=buffer, mode="wb") as gz:
            gz.write(data)
        data = buffer.getvalue()
        headers["Content-Encoding"] = "gzip"
    return data, headers


def _hash_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _file_url(file_id: str, variant: str) -> str:
    return f"/files/{file_id}/{variant}"


def _slides_json_url(file_id: str) -> str:
    return f"/files/{file_id}/slides.json"


def _slide_image_url(file_id: str, index: int) -> str:
    return f"/files/{file_id}/slide/{index}"


def _excel_summary_path(file_id: str) -> Path:
    return EXCEL_SUMMARY_DIR / f"{file_id}.json"


def _excel_summary_url(file_id: str) -> str:
    return f"/files/{file_id}/excel/summary.json"


def _excel_charts_json_url(file_id: str) -> str:
    return f"/files/{file_id}/excel/charts.json"


def _excel_chart_image_url(file_id: str, index: int) -> str:
    return f"/files/{file_id}/excel/chart/{index}"


def _excel_chart_sheets_json_path(file_id: str) -> Path:
    """Возвращает путь к JSON файлу с информацией о листах, содержащих диаграммы."""
    return EXCEL_CHARTS_META_DIR / f"{file_id}_sheets.json"


def _excel_chart_anchors_json_path(file_id: str) -> Path:
    """Возвращает путь к JSON файлу с якорями диаграмм."""
    return EXCEL_CHARTS_META_DIR / f"{file_id}_anchors.json"


def _get_code_file_path(asset: FileAsset) -> Path:
    path_str = asset.path_code_original or asset.path_original
    if not path_str:
        raise HTTPException(status_code=404, detail="Code file not available")
    path = Path(path_str)
    if not path.exists():
        raise HTTPException(status_code=404, detail="Code file missing on disk")
    return path


def _get_markdown_file_path(asset: FileAsset) -> Path:
    path_str = asset.path_markdown_raw or asset.path_original
    if not path_str:
        raise HTTPException(status_code=404, detail="Markdown file not available")
    path = Path(path_str)
    if not path.exists():
        raise HTTPException(status_code=404, detail="Markdown file missing on disk")
    return path


def _parse_xlsx_chart_anchors(path: Path) -> dict:
    """Парсит якоря (anchors) диаграмм из XLSX файла через DrawingML.
    
    Возвращает:
    {
        "sheets": [
            {
                "name": "Лист1",
                "charts": [
                    {
                        "rel": "xl/charts/chart1.xml",
                        "anchor": {
                            "from": {"col": 1, "row": 2},
                            "to": {"col": 8, "row": 12}
                        }
                    }
                ]
            }
        ]
    }
    """
    logger = logging.getLogger(__name__)
    result = {
        "sheets": []
    }
    
    try:
        with zipfile.ZipFile(path, 'r') as zf:
            # Находим листы и их пути
            ns_main = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
                      'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
                      'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                      'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'}
            
            # Читаем workbook.xml
            workbook_xml = zf.read('xl/workbook.xml')
            workbook_root = ET.fromstring(workbook_xml)
            
            sheets_info = []
            workbook_rels_xml = zf.read('xl/_rels/workbook.xml.rels')
            workbook_rels_root = ET.fromstring(workbook_rels_xml)
            
            for sheet in workbook_root.findall('.//main:sheet', ns_main):
                sheet_name = sheet.get('name')
                r_id = sheet.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                if sheet_name and r_id:
                    # Находим путь к листу
                    sheet_path = None
                    for rel in workbook_rels_root.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                        if rel.get('Id') == r_id:
                            sheet_path = rel.get('Target')
                            if not sheet_path.startswith('xl/'):
                                sheet_path = 'xl/' + sheet_path
                            break
                    
                    if sheet_path:
                        sheets_info.append((sheet_name, sheet_path))
            
            logger.info(f"Parsing chart anchors from {len(sheets_info)} sheet(s)")
            
            # Для каждого листа парсим якоря
            for sheet_name, sheet_path in sheets_info:
                try:
                    # Читаем sheet*.xml
                    sheet_xml = zf.read(sheet_path)
                    sheet_root = ET.fromstring(sheet_xml)
                    
                    # Ищем <drawing r:id>
                    drawing_ids = []
                    for drawing in sheet_root.findall('.//main:drawing', ns_main):
                        drawing_r_id = drawing.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                        if drawing_r_id:
                            drawing_ids.append(drawing_r_id)
                    
                    if not drawing_ids:
                        continue
                    
                    # Разрешаем drawing через relationships
                    sheet_dir = sheet_path.rsplit('/', 1)[0]  # xl/worksheets
                    rels_path = f'{sheet_dir}/_rels/{sheet_path.split("/")[-1]}.rels'
                    
                    try:
                        sheet_rels_xml = zf.read(rels_path)
                        sheet_rels_root = ET.fromstring(sheet_rels_xml)
                        
                        charts_on_sheet = []
                        
                        # Находим пути к drawing*.xml
                        for drawing_r_id in drawing_ids:
                            drawing_path = None
                            for rel in sheet_rels_root.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                                if rel.get('Id') == drawing_r_id:
                                    target = rel.get('Target')
                                    if not target.startswith('xl/'):
                                        target = f'{sheet_dir.rsplit("/", 1)[0]}/{target.lstrip("../")}'
                                    if not target.startswith('xl/'):
                                        target = 'xl/' + target.split('xl/', 1)[-1]
                                    drawing_path = target
                                    break
                            
                            if not drawing_path:
                                continue
                            
                            try:
                                # Читаем drawing*.xml
                                drawing_xml = zf.read(drawing_path)
                                drawing_root = ET.fromstring(drawing_xml)
                                
                                # Находим все twoCellAnchor и oneCellAnchor
                                for anchor in drawing_root.findall('.//xdr:twoCellAnchor', ns_main):
                                    chart_rel = None
                                    from_cell = None
                                    to_cell = None
                                    
                                    # Ищем связь с диаграммой
                                    for rel in anchor.findall('.//{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed', ns_main):
                                        chart_rel = rel.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                                        break
                                    
                                    if not chart_rel:
                                        continue
                                    
                                    # Разрешаем chart_rel через drawing*.xml.rels
                                    drawing_dir = drawing_path.rsplit('/', 1)[0]
                                    drawing_rels_path = f'{drawing_dir}/_rels/{drawing_path.split("/")[-1]}.rels'
                                    try:
                                        drawing_rels_xml = zf.read(drawing_rels_path)
                                        drawing_rels_root = ET.fromstring(drawing_rels_xml)
                                        
                                        chart_path = None
                                        for rel in drawing_rels_root.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                                            if rel.get('Id') == chart_rel:
                                                chart_path = rel.get('Target')
                                                break
                                        
                                        if not chart_path:
                                            continue
                                    except KeyError:
                                        continue
                                    
                                    # Парсим from/to ячейки
                                    from_elem = anchor.find('.//xdr:from', ns_main)
                                    to_elem = anchor.find('.//xdr:to', ns_main)
                                    
                                    if from_elem is not None and to_elem is not None:
                                        from_col_elem = from_elem.find('.//xdr:col', ns_main)
                                        from_row_elem = from_elem.find('.//xdr:row', ns_main)
                                        to_col_elem = to_elem.find('.//xdr:col', ns_main)
                                        to_row_elem = to_elem.find('.//xdr:row', ns_main)
                                        
                                        if (from_col_elem is not None and from_row_elem is not None and
                                            to_col_elem is not None and to_row_elem is not None):
                                            try:
                                                from_col = int(from_col_elem.text or 0)
                                                from_row = int(from_row_elem.text or 0)
                                                to_col = int(to_col_elem.text or 0)
                                                to_row = int(to_row_elem.text or 0)
                                                
                                                charts_on_sheet.append({
                                                    "rel": chart_path,
                                                    "anchor": {
                                                        "from": {"col": from_col, "row": from_row},
                                                        "to": {"col": to_col, "row": to_row}
                                                    }
                                                })
                                                logger.debug(f"Found chart anchor on '{sheet_name}': from=({from_col},{from_row}), to=({to_col},{to_row})")
                                            except (ValueError, TypeError):
                                                logger.debug(f"Invalid anchor cell values on '{sheet_name}'")
                                                continue
                                
                                # Также обрабатываем oneCellAnchor (упрощенно)
                                for anchor in drawing_root.findall('.//xdr:oneCellAnchor', ns_main):
                                    chart_rel = None
                                    for rel in anchor.findall('.//{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed', ns_main):
                                        chart_rel = rel.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                                        break
                                    
                                    if not chart_rel:
                                        continue
                                    
                                    # Разрешаем chart_rel через drawing*.xml.rels (как для twoCellAnchor)
                                    try:
                                        drawing_dir = drawing_path.rsplit('/', 1)[0]
                                        drawing_rels_path = f'{drawing_dir}/_rels/{drawing_path.split("/")[-1]}.rels'
                                        drawing_rels_xml = zf.read(drawing_rels_path)
                                        drawing_rels_root = ET.fromstring(drawing_rels_xml)
                                        
                                        chart_path = None
                                        for rel in drawing_rels_root.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                                            if rel.get('Id') == chart_rel:
                                                chart_path = rel.get('Target')
                                                break
                                        
                                        if not chart_path:
                                            continue
                                    except KeyError:
                                        continue
                                    
                                    # Для oneCellAnchor используем размеры из ext (EMU), конвертируем приблизительно
                                    from_elem = anchor.find('.//xdr:from', ns_main)
                                    ext_elem = anchor.find('.//xdr:ext', ns_main)
                                    
                                    if from_elem is not None and ext_elem is not None:
                                        from_col_elem = from_elem.find('.//xdr:col', ns_main)
                                        from_row_elem = from_elem.find('.//xdr:row', ns_main)
                                        cx = int(ext_elem.get('cx', 0))
                                        cy = int(ext_elem.get('cy', 0))
                                        
                                        if from_col_elem is not None and from_row_elem is not None and cx > 0 and cy > 0:
                                            try:
                                                from_col = int(from_col_elem.text or 0)
                                                from_row = int(from_row_elem.text or 0)
                                                
                                                # Конвертируем EMU в приблизительные колонки/строки
                                                # 1 колонка ~ 8.43 символов * 7 px ≈ 59 px, но в EMU это ~ 9525
                                                # 1 строка ~ 20 px, в EMU это ~ 19050
                                                col_span = max(1, int(cx / 9525))  # приблизительно
                                                row_span = max(1, int(cy / 19050))  # приблизительно
                                                
                                                charts_on_sheet.append({
                                                    "rel": chart_path,
                                                    "anchor": {
                                                        "from": {"col": from_col, "row": from_row},
                                                        "to": {"col": from_col + col_span, "row": from_row + row_span}
                                                    }
                                                })
                                                logger.debug(f"Found oneCellAnchor on '{sheet_name}': from=({from_col},{from_row}), span=({col_span},{row_span})")
                                            except (ValueError, TypeError):
                                                continue
                                
                            except KeyError:
                                logger.debug(f"Drawing XML not found: {drawing_path}")
                                continue
                            except Exception as e:
                                logger.debug(f"Error parsing drawing {drawing_path}: {e}")
                                continue
                        
                        if charts_on_sheet:
                            result["sheets"].append({
                                "name": sheet_name,
                                "charts": charts_on_sheet
                            })
                            logger.info(f"✅ Sheet '{sheet_name}' has {len(charts_on_sheet)} chart anchor(s)")
                    
                    except KeyError:
                        logger.debug(f"Sheet relationships not found: {rels_path}")
                        continue
                    except Exception as e:
                        logger.debug(f"Error processing sheet '{sheet_name}': {e}")
                        continue
                
                except KeyError:
                    logger.debug(f"Sheet XML not found: {sheet_path}")
                    continue
                except Exception as e:
                    logger.debug(f"Error processing sheet '{sheet_name}': {e}")
                    continue
    
    except Exception as e:
        logger.warning(f"Failed to parse chart anchors: {e}", exc_info=True)
        return result
    
    return result


def _detect_xlsx_charts_structural(path: Path) -> dict:
    """Структурно обнаруживает диаграммы в XLSX через парсинг XML.
    
    Алгоритм:
    1. Открывает XLSX как zip
    2. Парсит xl/workbook.xml для получения имен листов
    3. Для каждого листа: парсит xl/worksheets/sheet*.xml, ищет <drawing r:id>
    4. Разрешает drawing через xl/worksheets/_rels/sheet*.xml.rels
    5. В xl/drawings/drawing*.xml ищет связи с диаграммами через _rels
    6. Собирает маппинг: {sheetName: [chartPartIds...]}
    
    Возвращает:
    {
        "hasCharts": bool,
        "sheets": [
            {"name": str, "charts": [str]},
            ...
        ]
    }
    """
    logger = logging.getLogger(__name__)
    result = {
        "hasCharts": False,
        "sheets": []
    }
    
    try:
        with zipfile.ZipFile(path, 'r') as zf:
            # ШАГ 1: Получаем список листов из workbook.xml
            try:
                workbook_xml = zf.read('xl/workbook.xml')
                workbook_root = ET.fromstring(workbook_xml)
                
                # Находим имена листов
                ns = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
                      'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'}
                
                sheets_info = []
                for sheet in workbook_root.findall('.//main:sheet', ns):
                    sheet_name = sheet.get('name')
                    sheet_id = sheet.get('sheetId')
                    r_id = sheet.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                    if sheet_name and r_id:
                        sheets_info.append({
                            'name': sheet_name,
                            'sheet_id': sheet_id,
                            'r_id': r_id
                        })
                
                logger.info(f"Found {len(sheets_info)} sheet(s) in workbook")
                
            except Exception as e:
                logger.warning(f"Failed to parse workbook.xml: {e}")
                return result
            
            # ШАГ 2: Для каждого листа находим drawings и диаграммы
            for sheet_info in sheets_info:
                sheet_name = sheet_info['name']
                r_id = sheet_info['r_id']
                
                # Получаем путь к XML листа
                try:
                    # Разрешаем relationship для листа
                    rels_path = f'xl/_rels/workbook.xml.rels'
                    workbook_rels_xml = zf.read(rels_path)
                    workbook_rels_root = ET.fromstring(workbook_rels_xml)
                    
                    sheet_path = None
                    for rel in workbook_rels_root.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                        if rel.get('Id') == r_id:
                            sheet_path = rel.get('Target')
                            break
                    
                    if not sheet_path:
                        logger.debug(f"Could not find path for sheet '{sheet_name}'")
                        continue
                    
                    # Читаем sheet*.xml
                    if not sheet_path.startswith('xl/'):
                        sheet_path = 'xl/' + sheet_path
                    
                    sheet_xml = zf.read(sheet_path)
                    sheet_root = ET.fromstring(sheet_xml)
                    
                    # Ищем <drawing r:id>
                    drawings = []
                    for drawing in sheet_root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}drawing', ns):
                        drawing_r_id = drawing.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                        if drawing_r_id:
                            drawings.append(drawing_r_id)
                    
                    if not drawings:
                        logger.debug(f"No drawings found on sheet '{sheet_name}'")
                        continue
                    
                    logger.info(f"Found {len(drawings)} drawing(s) on sheet '{sheet_name}'")
                    
                    # ШАГ 3: Разрешаем drawings через relationships
                    sheet_dir = sheet_path.rsplit('/', 1)[0]  # xl/worksheets
                    rels_path = f'{sheet_dir}/_rels/{sheet_path.split("/")[-1]}.rels'
                    
                    charts_on_sheet = []
                    
                    try:
                        sheet_rels_xml = zf.read(rels_path)
                        sheet_rels_root = ET.fromstring(sheet_rels_xml)
                        
                        # Находим пути к drawing*.xml
                        drawing_paths = []
                        for rel in sheet_rels_root.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                            rel_id = rel.get('Id')
                            if rel_id in drawings:
                                target = rel.get('Target')
                                # Преобразуем относительный путь в абсолютный
                                if not target.startswith('xl/'):
                                    target = f'{sheet_dir.rsplit("/", 1)[0]}/{target.lstrip("../")}'
                                if not target.startswith('xl/'):
                                    target = 'xl/' + target.split('xl/', 1)[-1]
                                drawing_paths.append((rel_id, target))
                        
                        # ШАГ 4: В каждом drawing*.xml ищем связи с диаграммами
                        for rel_id, drawing_path in drawing_paths:
                            try:
                                # Парсим drawing*.xml.rels
                                drawing_dir = drawing_path.rsplit('/', 1)[0]
                                drawing_rels_path = f'{drawing_dir}/_rels/{drawing_path.split("/")[-1]}.rels'
                                
                                drawing_rels_xml = zf.read(drawing_rels_path)
                                drawing_rels_root = ET.fromstring(drawing_rels_xml)
                                
                                for rel in drawing_rels_root.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                                    rel_type = rel.get('Type', '')
                                    if 'chart' in rel_type.lower() or rel_type.endswith('/chart'):
                                        chart_part_id = rel.get('Target')
                                        charts_on_sheet.append(chart_part_id)
                                        logger.debug(f"Found chart in drawing: {chart_part_id}")
                                        
                            except KeyError:
                                logger.debug(f"Drawing relationships not found: {drawing_rels_path}")
                                continue
                            except Exception as e:
                                logger.debug(f"Error parsing drawing {drawing_path}: {e}")
                                continue
                        
                    except KeyError:
                        logger.debug(f"Sheet relationships not found: {rels_path}")
                        continue
                    except Exception as e:
                        logger.debug(f"Error parsing sheet relationships: {e}")
                        continue
                    
                    if charts_on_sheet:
                        result["sheets"].append({
                            "name": sheet_name,
                            "charts": charts_on_sheet
                        })
                        logger.info(f"✅ Sheet '{sheet_name}' has {len(charts_on_sheet)} chart(s)")
                    
                except KeyError as e:
                    logger.debug(f"Sheet XML not found for '{sheet_name}': {e}")
                    continue
                except Exception as e:
                    logger.debug(f"Error processing sheet '{sheet_name}': {e}")
                    continue
        
        result["hasCharts"] = len(result["sheets"]) > 0
        logger.info(f"Structural chart detection: hasCharts={result['hasCharts']}, sheets with charts: {len(result['sheets'])}")
        
    except Exception as e:
        logger.warning(f"Failed to detect charts structurally in {path}: {e}", exc_info=True)
        return result
    
    return result


def _cell_anchor_to_bbox(anchor, ws) -> Optional[Tuple[int, int, int, int]]:
    """Конвертирует якорь диаграммы (TwoCellAnchor/OneCellAnchor) в пиксельные координаты (x, y, width, height).
    
    Возвращает (x, y, w, h) в пикселях или None, если якорь невалиден.
    """
    try:
        from openpyxl.chart.shapes import TwoCellAnchor, OneCellAnchor
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None
    
    # Excel default размеры
    DEFAULT_COL_WIDTH = 8.43  # символы (Calibri 11)
    DEFAULT_ROW_HEIGHT = 15.0  # пункты
    
    # Конвертация в пиксели
    # Ширина колонки: ~ (width_chars * 7) + padding
    # Высота строки: points * 96/72 (96 DPI)
    CHAR_TO_PX = 7
    PADDING = 6
    PT_TO_PX = 96.0 / 72.0
    
    if isinstance(anchor, TwoCellAnchor):
        from_col = anchor.from_col
        from_row = anchor.from_row
        to_col = anchor.to_col
        to_row = anchor.to_row
    elif isinstance(anchor, OneCellAnchor):
        from_col = anchor.col
        from_row = anchor.row
        # OneCellAnchor - один ячейка, нужно взять примерные размеры
        to_col = from_col + 5  # примерная ширина диаграммы
        to_row = from_row + 10  # примерная высота диаграммы
    else:
        return None
    
    # Получаем defaultColWidth листа, если есть
    default_col_width = getattr(ws.sheet_format, 'defaultColWidth', None) or DEFAULT_COL_WIDTH
    
    # Считаем X координату: суммируем ширины всех колонок до from_col
    x = 0
    for col_idx in range(0, from_col):
        col_letter = get_column_letter(col_idx + 1)
        if col_letter in ws.column_dimensions:
            width = ws.column_dimensions[col_letter].width or default_col_width
        else:
            width = default_col_width
        x += round((width * CHAR_TO_PX) + PADDING)
    
    # Считаем ширину диаграммы: суммируем ширины колонок от from_col до to_col
    w = 0
    for col_idx in range(from_col, to_col + 1):
        col_letter = get_column_letter(col_idx + 1)
        if col_letter in ws.column_dimensions:
            width = ws.column_dimensions[col_letter].width or default_col_width
        else:
            width = default_col_width
        w += round((width * CHAR_TO_PX) + PADDING)
    
    # Считаем Y координату: суммируем высоты всех строк до from_row
    y = 0
    for row_idx in range(0, from_row):
        row_num = row_idx + 1
        if row_num in ws.row_dimensions:
            height = ws.row_dimensions[row_num].height or DEFAULT_ROW_HEIGHT
        else:
            height = DEFAULT_ROW_HEIGHT
        y += round(height * PT_TO_PX)
    
    # Считаем высоту диаграммы: суммируем высоты строк от from_row до to_row
    h = 0
    for row_idx in range(from_row, to_row + 1):
        row_num = row_idx + 1
        if row_num in ws.row_dimensions:
            height = ws.row_dimensions[row_num].height or DEFAULT_ROW_HEIGHT
        else:
            height = DEFAULT_ROW_HEIGHT
        h += round(height * PT_TO_PX)
    
    return (x, y, w, h)


def _extract_xlsx_charts_precise(original_path: Path, file_id: str) -> Optional[tuple[Path, Path, int, Path]]:
    """Точно извлекает диаграммы из XLSX по якорям OpenXML.
    
    Алгоритм:
    1. Читаем диаграммы и их якоря через openpyxl
    2. Конвертируем якоря (col, row) в пиксельные координаты
    3. Рендерим каждый лист в PNG через LibreOffice
    4. Вырезаем диаграммы из изображений листов по координатам
    5. Сохраняем только диаграммы (без таблиц)
    
    Возвращает (preview_path, json_path, charts_count, charts_dir) или None.
    """
    logger = logging.getLogger(__name__)
    
    # openpyxl уже импортирован в начале файла
    
    # ШАГ 1: Открываем XLSX и находим все диаграммы с якорями
    logger.info(f"Reading XLSX file {file_id} to find chart anchors...")
    
    # read_only=False нужен для доступа к размерам колонок/строк
    try:
        wb = load_workbook(original_path, read_only=False, keep_links=False, data_only=True)
    except Exception as e:
        logger.warning(f"Failed to open XLSX file {file_id}: {e}")
        return None
    
    charts_data = []  # Список: (sheet_name, chart_idx, bbox, chart_type, title)
    
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Ищем диаграммы на листе - проверяем разные способы доступа
            charts_on_sheet = []
            
            # Способ 1: ws._charts (для встроенных диаграмм)
            if hasattr(ws, '_charts'):
                charts_list = ws._charts
                if charts_list:
                    logger.info(f"Found {len(charts_list)} chart(s) via _charts on sheet '{sheet_name}'")
                    for ch in charts_list:
                        logger.debug(f"  Chart type: {type(ch).__name__}, has anchor attr: {hasattr(ch, 'anchor')}")
                    charts_on_sheet.extend(charts_list)
                else:
                    logger.debug(f"No charts in ws._charts for sheet '{sheet_name}'")
            else:
                logger.debug(f"Sheet '{sheet_name}' has no _charts attribute")
            
            # Способ 2: ws._images (ChartObject может быть в _images)
            if hasattr(ws, '_images') and ws._images:
                logger.info(f"Found {len(ws._images)} image(s) on sheet '{sheet_name}'")
                for img in ws._images:
                    if hasattr(img, 'anchor'):
                        # Проверяем, является ли это диаграммой
                        logger.debug(f"Image object found: {type(img).__name__}, has anchor: {hasattr(img, 'anchor')}")
                        # Если это ChartObject, попробуем добавить его
                        try:
                            from openpyxl.chart.chartspace import ChartSpace
                            if hasattr(img, 'chart') or 'Chart' in type(img).__name__:
                                logger.info(f"Found ChartObject in _images: {type(img).__name__}")
                                # ChartObject имеет anchor, попробуем использовать его
                                if hasattr(img, 'anchor'):
                                    charts_on_sheet.append(img)
                        except Exception as e:
                            logger.debug(f"Error checking image for chart: {e}")
            
            # Способ 3: Проверяем через drawings (для ChartObject)
            try:
                if hasattr(ws, '_rels') and ws._rels:
                    for rel_id, rel in ws._rels.items():
                        if 'chart' in str(rel.target):
                            logger.debug(f"Found chart relationship: {rel.target}")
            except Exception as e:
                logger.debug(f"Error checking _rels: {e}")
            
            if not charts_on_sheet:
                logger.debug(f"No charts found on sheet '{sheet_name}'")
                continue
            
            logger.info(f"Processing {len(charts_on_sheet)} chart(s) on sheet '{sheet_name}'")
            
            for idx, chart in enumerate(charts_on_sheet):
                chart_type_name = type(chart).__name__
                logger.info(f"Processing chart {idx} on sheet '{sheet_name}': type={chart_type_name}")
                
                # Пытаемся получить якорь диаграммы
                anchor = None
                
                # Способ 1: chart.anchor (для Chart объекта или ChartObject)
                if hasattr(chart, 'anchor'):
                    try:
                        anchor = chart.anchor
                        logger.info(f"✅ Chart {idx}: found anchor via chart.anchor: {type(anchor).__name__}")
                    except Exception as e:
                        logger.debug(f"Error accessing chart.anchor: {e}")
                
                # Способ 2: chart.parent.anchor (если диаграмма вложена в ChartObject)
                if anchor is None and hasattr(chart, 'parent'):
                    try:
                        parent = chart.parent
                        if hasattr(parent, 'anchor'):
                            anchor = parent.anchor
                            logger.info(f"✅ Chart {idx}: found anchor via chart.parent.anchor: {type(anchor).__name__}")
                    except Exception as e:
                        logger.debug(f"Error accessing chart.parent.anchor: {e}")
                
                # Способ 3: Проверяем, является ли chart ChartObject напрямую
                if anchor is None:
                    try:
                        # В openpyxl ChartObject может иметь anchor напрямую
                        if 'Object' in chart_type_name or 'Image' in chart_type_name:
                            if hasattr(chart, '_anchor'):
                                anchor = chart._anchor
                                logger.info(f"✅ Chart {idx}: found anchor via chart._anchor: {type(anchor).__name__}")
                    except Exception as e:
                        logger.debug(f"Error checking _anchor: {e}")
                
                if anchor is None:
                    logger.warning(f"❌ Chart {idx} on sheet '{sheet_name}': cannot find anchor (type={chart_type_name})")
                    logger.warning(f"   Chart attributes: {[a for a in dir(chart) if not a.startswith('__') and 'anchor' in a.lower()]}")
                    continue
                
                # Конвертируем якорь в bbox
                bbox = _cell_anchor_to_bbox(anchor, ws)
                if bbox is None:
                    logger.warning(f"Chart {idx} on sheet '{sheet_name}': failed to convert anchor to bbox")
                    continue
                
                x, y, w, h = bbox
                chart_type = type(chart).__name__
                chart_title = getattr(chart, 'title', None)
                if chart_title:
                    title_text = chart_title.t if hasattr(chart_title, 't') else str(chart_title)
                else:
                    title_text = f"Chart {idx + 1}"
                
                charts_data.append({
                    'sheet': sheet_name,
                    'index': len(charts_data) + 1,  # Глобальный индекс диаграммы
                    'bbox': [x, y, w, h],
                    'chart_type': chart_type,
                    'title': title_text,
                })
                
                logger.info(f"✅ Chart {len(charts_data)}: sheet='{sheet_name}', type={chart_type}, bbox=({x},{y},{w},{h}), title='{title_text}'")
        
        wb.close()
        
    except Exception as e:
        logger.warning(f"Error reading chart anchors from {file_id}: {e}", exc_info=True)
        wb.close()
        return None
    
    if not charts_data:
        logger.warning(f"⚠️ No charts with valid anchors found in XLSX file {file_id}")
        logger.warning("   This may mean:")
        logger.warning("   1. File has no charts")
        logger.warning("   2. Charts don't have accessible anchors in openpyxl")
        logger.warning("   3. Charts use unsupported anchor format")
        return None
    
    logger.info(f"✅ Found {len(charts_data)} chart(s) with valid anchors in {file_id}")
    
    # ШАГ 2: Рендерим листы в изображения через LibreOffice
    soffice_path = _find_libreoffice()
    if not soffice_path:
        logger.warning(f"LibreOffice not found for {file_id}, cannot render sheets")
        return None
    
    # Конвертируем XLSX в PDF (каждый лист = страница)
    logger.info(f"Converting XLSX to PDF to render sheets...")
    try:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            cmd = [
                soffice_path,
                "--headless",
                "--convert-to",
                "pdf",
                "--outdir",
                str(tmp_path),
                str(original_path),
            ]
            result = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, check=True, timeout=60)
            pdf_path = tmp_path / f"{original_path.stem}.pdf"
            if not pdf_path.exists():
                logger.warning(f"PDF not generated for {file_id}")
                return None
            pdf_bytes = pdf_path.read_bytes()
    except Exception as e:
        logger.warning(f"Failed to convert XLSX to PDF for {file_id}: {e}")
        return None
    
    # Рендерим каждую страницу PDF как PNG (каждая страница = лист)
    pdf_pages = _pdf_page_count(pdf_bytes)
    if pdf_pages == 0:
        logger.warning(f"PDF has no pages for {file_id}")
        return None
    
    logger.info(f"PDF has {pdf_pages} pages (sheets), rendering as images...")
    
    # Маппинг: sheet_name -> page_number (предполагаем, что порядок листов = порядок страниц)
    # Сохраняем имена листов ДО закрытия книги
    sheet_names = list(wb.sheetnames)
    wb.close()  # Закрываем книгу, больше не нужна
    sheet_to_page = {name: idx + 1 for idx, name in enumerate(sheet_names)}
    
    # ШАГ 3: Вырезаем диаграммы из изображений листов
    charts_dir = _excel_charts_dir_path(file_id)
    charts_dir.mkdir(parents=True, exist_ok=True)
    
    rendered_charts = []
    preview_path: Optional[Path] = None
    
    for chart_data in charts_data:
        sheet_name = chart_data['sheet']
        page_num = sheet_to_page.get(sheet_name)
        
        if not page_num or page_num > pdf_pages:
            logger.warning(f"Sheet '{sheet_name}' not found in PDF pages for {file_id}")
            continue
        
        # Рендерим страницу PDF как PNG
        sheet_image_bytes = _render_pdf_page(pdf_bytes, page_num, scale=1.2)
        if not sheet_image_bytes:
            logger.warning(f"Failed to render page {page_num} (sheet '{sheet_name}') for {file_id}")
            continue
        
        # Открываем изображение листа
        try:
            sheet_image = Image.open(BytesIO(sheet_image_bytes))
            sheet_width, sheet_height = sheet_image.size
        except Exception as e:
            logger.warning(f"Failed to open sheet image for {file_id}: {e}")
            continue
        
        # Получаем bbox диаграммы
        x, y, w, h = chart_data['bbox']
        
        # Добавляем padding чтобы не обрезать оси/легенды (+10px со всех сторон)
        padding = 10
        x = max(0, x - padding)
        y = max(0, y - padding)
        w = min(sheet_width - x, w + padding * 2)
        h = min(sheet_height - y, h + padding * 2)
        
        # Вырезаем диаграмму из изображения листа
        try:
            chart_image = sheet_image.crop((x, y, x + w, y + h))
        except Exception as e:
            logger.warning(f"Failed to crop chart image for {file_id}: {e}")
            continue
        
        # Сохраняем как WebP
        chart_index = chart_data['index']
        chart_path = charts_dir / f"{chart_index}.webp"
        
        buffer = BytesIO()
        chart_image.save(buffer, format="WEBP", quality=85)
        _write_file(chart_path, buffer.getvalue())
        
        rendered_charts.append({
            'sheet': sheet_name,
            'index': chart_index,
            'src': _excel_chart_image_url(file_id, chart_index),
            'bbox': [x, y, w, h],
            'title': chart_data['title'],
        })
        
        if preview_path is None:
            preview_path = chart_path
        
        logger.info(f"Extracted chart {chart_index} from sheet '{sheet_name}' (page {page_num}), saved to {chart_path}")
    
    if not rendered_charts:
        logger.error(f"❌ Failed to extract any charts from {file_id}")
        logger.error(f"   Found {len(charts_data)} chart(s) with anchors, but failed to extract images")
        logger.error(f"   This may indicate problems with:")
        logger.error(f"   - PDF rendering")
        logger.error(f"   - Sheet-to-page mapping")
        logger.error(f"   - Bbox coordinates calculation")
        return None
    
    # ШАГ 4: Сохраняем метаданные
    meta = {
        "count": len(rendered_charts),
        "format": "webp",
        "charts": rendered_charts,
    }
    
    json_path = _excel_charts_json_path(file_id)
    json_path.write_text(json.dumps(meta, ensure_ascii=False), encoding="utf-8")
    
    logger.info(f"Successfully extracted {len(rendered_charts)} chart(s) from {file_id}")
    
    return preview_path, json_path, len(rendered_charts), charts_dir


def _find_charts_in_excel(path: Path) -> list[dict]:
    """Находит все диаграммы в Excel файле через openpyxl.
    
    Возвращает список словарей с информацией о диаграммах: {'sheet': str, 'chart_type': str, 'title': str}
    """
    logger = logging.getLogger(__name__)
    charts_info = []
    
    try:
        # Открываем файл в режиме чтения (read_only=False нужен для доступа к диаграммам)
        wb = load_workbook(path, read_only=False, keep_links=False)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Проверяем наличие диаграмм на листе
            # openpyxl хранит диаграммы в sheet._charts
            if hasattr(ws, '_charts') and ws._charts:
                for idx, chart in enumerate(ws._charts):
                    chart_type = type(chart).__name__
                    chart_title = getattr(chart, 'title', None)
                    if chart_title:
                        title_text = chart_title.t if hasattr(chart_title, 't') else str(chart_title)
                    else:
                        title_text = f"Chart {idx + 1}"
                    
                    charts_info.append({
                        'sheet': sheet_name,
                        'chart_index': idx,
                        'chart_type': chart_type,
                        'title': title_text,
                    })
                    logger.info(f"Found chart in sheet '{sheet_name}': {chart_type} - {title_text}")
        
        wb.close()
        
    except Exception as e:
        logger.warning(f"Error reading charts from Excel file {path}: {e}", exc_info=True)
    
    return charts_info


def _export_charts_via_libreoffice_uno(original_path: Path, file_id: str, charts_info: list[dict]) -> Optional[tuple[Path, Path, int, Path]]:
    """Экспортирует диаграммы напрямую через LibreOffice Python-UNO API.
    
    Это ПРЯМОЙ доступ к объектам диаграмм, без фильтрации PDF страниц.
    """
    logger = logging.getLogger(__name__)
    
    try:
        import uno
        from com.sun.star.beans import PropertyValue
        from com.sun.star.connection import NoConnectException
    except ImportError:
        logger.warning("LibreOffice Python-UNO not available - falling back to PDF filtering method")
        return None
    
    soffice_path = _find_libreoffice()
    if not soffice_path:
        logger.warning("LibreOffice not found - cannot use UNO API")
        return None
    
    # Запускаем LibreOffice в режиме для UNO
    # Это сложно, требует запуска soffice с параметрами для UNO сервера
    # Пока оставляем как заглушку - это требует дополнительной настройки
    logger.warning("LibreOffice UNO API export not yet implemented - using PDF filtering fallback")
    return None


def _compute_sheet_geometry(path: Path, sheet_name: str) -> dict:
    """Вычисляет геометрию листа (ширины колонок и высоты строк в пикселях).
    
    Возвращает:
    {
        "colWidthsPx": [64, 64, 80, ...],  # ширина каждой колонки в пикселях
        "rowHeightsPx": [20, 20, 25, ...],  # высота каждой строки в пикселях
        "colPrefixPx": [0, 64, 128, ...],  # префиксные суммы для колонок
        "rowPrefixPx": [0, 20, 40, ...]   # префиксные суммы для строк
    }
    """
    logger = logging.getLogger(__name__)
    
    # Значения по умолчанию Excel
    DEFAULT_COL_WIDTH_PX = 64  # ~8.43 символов * 7.6 px при 96 DPI
    DEFAULT_ROW_HEIGHT_PX = 20  # ~15 pt
    
    col_widths_px = []
    row_heights_px = []
    
    try:
        with zipfile.ZipFile(path, 'r') as zf:
            # Находим путь к листу
            workbook_xml = zf.read('xl/workbook.xml')
            workbook_root = ET.fromstring(workbook_xml)
            
            ns = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
                  'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'}
            
            sheet_path = None
            workbook_rels_xml = zf.read('xl/_rels/workbook.xml.rels')
            workbook_rels_root = ET.fromstring(workbook_rels_xml)
            
            for sheet in workbook_root.findall('.//main:sheet', ns):
                if sheet.get('name') == sheet_name:
                    r_id = sheet.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                    if r_id:
                        for rel in workbook_rels_root.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                            if rel.get('Id') == r_id:
                                sheet_path = rel.get('Target')
                                if not sheet_path.startswith('xl/'):
                                    sheet_path = 'xl/' + sheet_path
                                break
                    break
            
            if not sheet_path:
                logger.warning(f"Sheet '{sheet_name}' not found for geometry computation")
                return {"colWidthsPx": [], "rowHeightsPx": [], "colPrefixPx": [0], "rowPrefixPx": [0]}
            
            # Читаем sheet*.xml
            sheet_xml = zf.read(sheet_path)
            sheet_root = ET.fromstring(sheet_xml)
            
            # Парсим ширины колонок из <cols><col width=...>
            max_col = 0
            col_widths_map = {}
            for col in sheet_root.findall('.//main:col', ns):
                min_attr = col.get('min')
                max_attr = col.get('max')
                width_attr = col.get('width')
                
                if min_attr and max_attr and width_attr:
                    try:
                        min_col = int(min_attr) - 1  # Excel uses 1-based, we use 0-based
                        max_col_idx = int(max_attr) - 1
                        # Excel width is in characters, convert to pixels
                        width_chars = float(width_attr)
                        width_px = int(width_chars * 7.6)  # approximate conversion
                        
                        for col_idx in range(min_col, max_col_idx + 1):
                            col_widths_map[col_idx] = width_px
                            max_col = max(max_col, col_idx)
                    except (ValueError, TypeError):
                        continue
            
            # Парсим высоты строк из <row ht=...>
            max_row = 0
            row_heights_map = {}
            for row in sheet_root.findall('.//main:row', ns):
                r_attr = row.get('r')
                ht_attr = row.get('ht')
                
                if r_attr and ht_attr:
                    try:
                        row_idx = int(r_attr) - 1  # Excel uses 1-based, we use 0-based
                        # Excel height is in points, convert to pixels
                        height_pt = float(ht_attr)
                        height_px = int(height_pt * 1.33)  # 1 pt ≈ 1.33 px at 96 DPI
                        
                        row_heights_map[row_idx] = height_px
                        max_row = max(max_row, row_idx)
                    except (ValueError, TypeError):
                        continue
            
            # Строим массивы с дефолтными значениями
            # Используем разумный максимум: до 100 колонок и 1000 строк
            max_col = max(max_col, 99)
            max_row = max(max_row, 999)
            
            col_widths_px = [col_widths_map.get(i, DEFAULT_COL_WIDTH_PX) for i in range(max_col + 1)]
            row_heights_px = [row_heights_map.get(i, DEFAULT_ROW_HEIGHT_PX) for i in range(max_row + 1)]
            
            # Строим префиксные суммы
            col_prefix_px = [0]
            for w in col_widths_px:
                col_prefix_px.append(col_prefix_px[-1] + w)
            
            row_prefix_px = [0]
            for h in row_heights_px:
                row_prefix_px.append(row_prefix_px[-1] + h)
            
            logger.debug(f"Computed geometry for '{sheet_name}': {len(col_widths_px)} cols, {len(row_heights_px)} rows")
            
    except Exception as e:
        logger.warning(f"Failed to compute sheet geometry for '{sheet_name}': {e}")
        # Возвращаем дефолты
        col_widths_px = [DEFAULT_COL_WIDTH_PX] * 100
        row_heights_px = [DEFAULT_ROW_HEIGHT_PX] * 1000
        
        col_prefix_px = [0] + [DEFAULT_COL_WIDTH_PX * i for i in range(1, 101)]
        row_prefix_px = [0] + [DEFAULT_ROW_HEIGHT_PX * i for i in range(1, 1001)]
    
    return {
        "colWidthsPx": col_widths_px,
        "rowHeightsPx": row_heights_px,
        "colPrefixPx": col_prefix_px,
        "rowPrefixPx": row_prefix_px
    }


def _build_sheet_to_pages_mapping(pdf_bytes: bytes, wb) -> dict:
    """Строит маппинг листов Excel на страницы PDF.
    
    Возвращает:
    {
        "sheets": [
            {"sheet": "Лист1", "pages": [1]},  # индексы страниц (1-based)
            ...
        ]
    }
    
    Примечание: LibreOffice обычно создает одну страницу на лист,
    но некоторые листы могут занимать несколько страниц.
    Простейшая эвристика: порядок листов = порядок страниц, по одной странице на лист.
    """
    logger = logging.getLogger(__name__)
    pdf_pages = _pdf_page_count(pdf_bytes)
    sheet_names = list(wb.sheetnames)
    
    # Простейшая эвристика: 1 лист = 1 страница (если страниц столько же, сколько листов)
    mapping = []
    if pdf_pages == len(sheet_names):
        # Точное соответствие: каждый лист = одна страница
        for idx, sheet_name in enumerate(sheet_names):
            mapping.append({"sheet": sheet_name, "pages": [idx + 1]})
        logger.info(f"Built sheet-to-pages mapping: {len(sheet_names)} sheets = {pdf_pages} pages (1:1)")
    else:
        # Неточное соответствие: распределяем страницы равномерно
        pages_per_sheet = pdf_pages // len(sheet_names) if sheet_names else 1
        remainder = pdf_pages % len(sheet_names)
        page_idx = 1
        
        for idx, sheet_name in enumerate(sheet_names):
            pages = pages_per_sheet
            if idx < remainder:
                pages += 1
            sheet_pages = list(range(page_idx, page_idx + pages))
            mapping.append({"sheet": sheet_name, "pages": sheet_pages})
            page_idx += pages
        logger.info(f"Built sheet-to-pages mapping: {len(sheet_names)} sheets = {pdf_pages} pages (distributed)")
    
    return {"sheets": mapping}


def _convert_excel_to_charts(original_path: Path, file_id: str) -> Optional[tuple[Path, Path, int, Path]]:
    """Извлекает диаграммы из Excel файла, вырезая только области диаграмм по якорям.
    
    НОВАЯ АРХИТЕКТУРА (точное вырезание по якорям):
    1. Парсим якоря диаграмм через DrawingML (_parse_xlsx_chart_anchors)
    2. Сохраняем якоря в JSON файл
    3. Для каждого листа с диаграммами:
       - Вычисляем геометрию листа (ширины колонок, высоты строк)
       - Рендерим лист в изображение через LibreOffice
       - Вырезаем области диаграмм по якорям
       - Сохраняем только вырезанные области
    4. Если якорей нет или LibreOffice недоступен - возвращаем None (graceful fallback)
    
    Возвращает (preview_path, json_path, charts_count, charts_dir) или None.
    """
    logger = logging.getLogger(__name__)
    
    # Определяем формат файла
    file_ext = original_path.suffix.lower()
    
    if file_ext != '.xlsx':
        # Для .xls/.xlsm пока не поддерживаем диаграммы
        logger.info(f"Chart extraction not yet supported for {file_ext} format (file {file_id})")
        return None
    
    # ШАГ 1: Парсим якоря диаграмм через DrawingML
    logger.info(f"🔍 Parsing chart anchors from XLSX file {file_id}...")
    anchors_data = _parse_xlsx_chart_anchors(original_path)
    
    if not anchors_data.get("sheets") or not any(s.get("charts") for s in anchors_data["sheets"]):
        logger.info(f"No chart anchors found in {file_id}")
        # Сохраняем пустые якоря для отладки
        anchors_json_path = _excel_chart_anchors_json_path(file_id)
        anchors_json_path.write_text(json.dumps(anchors_data, ensure_ascii=False), encoding="utf-8")
        return None
    
    total_charts = sum(len(s.get("charts", [])) for s in anchors_data["sheets"])
    logger.info(f"✅ Found {total_charts} chart anchor(s) across {len(anchors_data['sheets'])} sheet(s)")
    
    # Сохраняем якоря
    anchors_json_path = _excel_chart_anchors_json_path(file_id)
    anchors_json_path.write_text(json.dumps(anchors_data, ensure_ascii=False), encoding="utf-8")
    logger.info(f"Saved chart anchors to {anchors_json_path}")
    
    # ШАГ 2: Проверяем наличие LibreOffice
    soffice_path = _find_libreoffice()
    if not soffice_path:
        logger.warning(f"LibreOffice not found for {file_id}, cannot render sheets for cropping")
        return None
    
    # ШАГ 3: Рендерим каждый лист и вырезаем диаграммы
    charts_dir = _excel_charts_dir_path(file_id)
    charts_dir.mkdir(parents=True, exist_ok=True)
    
    rendered_charts = []
    preview_path: Optional[Path] = None
    width = height = None
    chart_index = 1
    
    # Открываем XLSX через openpyxl для получения имен листов (для маппинга на страницы PDF)
    try:
        wb = load_workbook(original_path, read_only=True, data_only=True)
        sheet_names = list(wb.sheetnames)
        wb.close()
    except Exception as e:
        logger.warning(f"Failed to read workbook for sheet names: {e}")
        sheet_names = [s["name"] for s in anchors_data["sheets"]]
    
    # Конвертируем XLSX в PDF (один раз для всех листов)
    logger.info(f"Converting XLSX to PDF to render sheets...")
    try:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            cmd = [
                soffice_path,
                "--headless",
                "--convert-to",
                "pdf",
                "--outdir",
                str(tmp_path),
                str(original_path),
            ]
            result = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, check=True, timeout=60)
            pdf_path = tmp_path / f"{original_path.stem}.pdf"
            if not pdf_path.exists():
                logger.warning(f"PDF not generated for {file_id}")
                return None
            pdf_bytes = pdf_path.read_bytes()
    except Exception as e:
        logger.warning(f"Failed to convert XLSX to PDF for {file_id}: {e}")
        return None
    
    pdf_pages = _pdf_page_count(pdf_bytes)
    if pdf_pages == 0:
        logger.warning(f"PDF has no pages for {file_id}")
        return None
    
    # Строим маппинг листов на страницы (предполагаем 1:1)
    sheet_to_page = {name: idx + 1 for idx, name in enumerate(sheet_names) if idx < pdf_pages}
    
    # ШАГ 4: Для каждого листа с диаграммами вырезаем области
    for sheet_data in anchors_data["sheets"]:
        sheet_name = sheet_data["name"]
        charts = sheet_data.get("charts", [])
        
        if not charts:
            continue
        
        # Получаем номер страницы для этого листа
        page_num = sheet_to_page.get(sheet_name)
        if not page_num:
            logger.warning(f"Sheet '{sheet_name}' not found in PDF pages")
            continue
        
        # Рендерим страницу как изображение
        logger.info(f"Rendering page {page_num} (sheet '{sheet_name}') for cropping charts...")
        sheet_image_bytes = _render_pdf_page(pdf_bytes, page_num, scale=1.2)
        if not sheet_image_bytes:
            logger.warning(f"Failed to render page {page_num} for sheet '{sheet_name}'")
            continue
        
        sheet_image = Image.open(BytesIO(sheet_image_bytes))
        sheet_width, sheet_height = sheet_image.size
        
        # Вычисляем геометрию листа
        geometry = _compute_sheet_geometry(original_path, sheet_name)
        col_prefix = geometry.get("colPrefixPx", [0])
        row_prefix = geometry.get("rowPrefixPx", [0])
        
        # Для каждой диаграммы на листе вырезаем область
        for chart_data in charts:
            anchor = chart_data.get("anchor", {})
            from_cell = anchor.get("from", {})
            to_cell = anchor.get("to", {})
            
            if not from_cell or not to_cell:
                logger.warning(f"Invalid anchor for chart on sheet '{sheet_name}'")
                continue
            
            try:
                from_col = from_cell.get("col", 0)
                from_row = from_cell.get("row", 0)
                to_col = to_cell.get("col", 0)
                to_row = to_cell.get("row", 0)
                
                # Вычисляем пиксельные координаты
                if from_col >= len(col_prefix) or to_col >= len(col_prefix):
                    logger.warning(f"Column index out of bounds: col={from_col}-{to_col}, max={len(col_prefix)-1}")
                    continue
                if from_row >= len(row_prefix) or to_row >= len(row_prefix):
                    logger.warning(f"Row index out of bounds: row={from_row}-{to_row}, max={len(row_prefix)-1}")
                    continue
                
                x1 = col_prefix[from_col]
                y1 = row_prefix[from_row]
                x2 = col_prefix[to_col] if to_col < len(col_prefix) else col_prefix[-1]
                y2 = row_prefix[to_row] if to_row < len(row_prefix) else row_prefix[-1]
                
                # Добавляем padding
                padding = 12
                x1 = max(0, x1 - padding)
                y1 = max(0, y1 - padding)
                x2 = min(sheet_width, x2 + padding)
                y2 = min(sheet_height, y2 + padding)
                
                # Проверяем, что область валидна
                if x2 <= x1 or y2 <= y1:
                    logger.warning(f"Invalid crop area: ({x1},{y1}) to ({x2},{y2})")
                    continue
                
                # Вырезаем область
                try:
                    cropped_image = sheet_image.crop((x1, y1, x2, y2))
                except Exception as e:
                    logger.warning(f"Failed to crop chart area: {e}")
                    continue
                
                # Масштабируем, если нужно
                if EXCEL_CHARTS_TARGET_WIDTH and cropped_image.width > EXCEL_CHARTS_TARGET_WIDTH:
                    cropped_image.thumbnail((EXCEL_CHARTS_TARGET_WIDTH, EXCEL_CHARTS_TARGET_WIDTH * 2), Image.Resampling.LANCZOS)
                
                # Сохраняем как WebP
                buffer = BytesIO()
                cropped_image.save(buffer, format="WEBP", quality=85)
                chart_bytes = buffer.getvalue()
                
                chart_path = charts_dir / f"{chart_index}.webp"
                _write_file(chart_path, chart_bytes)
                
                rendered_charts.append({
                    'sheet': sheet_name,
                    'index': chart_index,
                    'src': _excel_chart_image_url(file_id, chart_index),
                    'anchor': anchor,
                })
                
                if preview_path is None:
                    preview_path = chart_path
                    width, height = cropped_image.size
                
                logger.info(f"✅ Cropped chart {chart_index} from sheet '{sheet_name}': ({x1},{y1})-({x2},{y2}), size={cropped_image.width}x{cropped_image.height}")
                chart_index += 1
                
            except Exception as e:
                logger.warning(f"Error processing chart anchor on sheet '{sheet_name}': {e}")
                continue
    
    if not rendered_charts:
        logger.warning(f"Failed to crop any charts from {file_id}")
        return None
    
    # ШАГ 5: Сохраняем метаданные
    meta = {
        "count": len(rendered_charts),
        "format": "webp",
        "width": width,
        "height": height,
        "mode": "anchors",  # Используем режим вырезания по якорям
        "charts": rendered_charts,
    }
    
    json_path = _excel_charts_json_path(file_id)
    json_path.write_text(json.dumps(meta, ensure_ascii=False), encoding="utf-8")
    
    logger.info(f"✅ Successfully extracted {len(rendered_charts)} chart(s) from {file_id} using anchor-based cropping")
    
    return preview_path, json_path, len(rendered_charts), charts_dir


def _crop_chart_area(image: Image.Image) -> Image.Image:
    """Обрезает изображение, оставляя только область с контентом (диаграммой), убирая пустые поля.
    
    Пытается найти область с контентом, определяя края, которые не являются чисто белыми.
    """
    # Конвертируем в RGB если нужно
    if image.mode != 'RGB':
        image = image.convert('RGB')
    
    width, height = image.size
    pixels = image.load()
    
    # Порог для определения "не белого" пикселя (допустим небольшие отклонения от белого)
    threshold = 240  # RGB значение меньше 240 считается не белым
    
    # Находим левую границу
    left = 0
    for x in range(width):
        for y in range(height):
            r, g, b = pixels[x, y]
            if r < threshold or g < threshold or b < threshold:
                left = x
                break
        if left > 0:
            break
    else:
        # Если весь файл белый, возвращаем оригинал
        return image
    
    # Находим правую границу
    right = width - 1
    for x in range(width - 1, left - 1, -1):
        for y in range(height):
            r, g, b = pixels[x, y]
            if r < threshold or g < threshold or b < threshold:
                right = x
                break
        if right < width - 1:
            break
    
    # Находим верхнюю границу
    top = 0
    for y in range(height):
        for x in range(left, right + 1):
            r, g, b = pixels[x, y]
            if r < threshold or g < threshold or b < threshold:
                top = y
                break
        if top > 0:
            break
    
    # Находим нижнюю границу
    bottom = height - 1
    for y in range(height - 1, top - 1, -1):
        for x in range(left, right + 1):
            r, g, b = pixels[x, y]
            if r < threshold or g < threshold or b < threshold:
                bottom = y
                break
        if bottom < height - 1:
            break
    
    # Обрезаем до найденных границ
    if left < right and top < bottom:
        cropped = image.crop((left, top, right + 1, bottom + 1))
        # Добавляем небольшой отступ (5% от размера)
        padding = max(cropped.width, cropped.height) // 20
        # Создаем новое изображение с отступом на белом фоне
        new_width = cropped.width + padding * 2
        new_height = cropped.height + padding * 2
        new_image = Image.new('RGB', (new_width, new_height), color=(255, 255, 255))
        new_image.paste(cropped, (padding, padding))
        return new_image
    
    return image


def _detect_table_structure(image: Image.Image) -> tuple[float, float]:
    """Обнаруживает признаки табличной структуры: регулярные горизонтальные и вертикальные линии.
    
    Возвращает (horizontal_lines_score, vertical_lines_score) - оценки наличия регулярных линий.
    """
    width, height = image.size
    pixels = image.load()
    threshold = 200  # Порог для черных линий (таблицы обычно имеют темные линии)
    
    # Проверяем горизонтальные линии (каждую 10-ю строку для скорости)
    horizontal_step = max(1, height // 50)
    vertical_step = max(1, width // 50)
    
    horizontal_line_count = 0
    vertical_line_count = 0
    
    # Ищем горизонтальные линии (проходят через большую часть ширины)
    for y in range(0, height, horizontal_step):
        line_pixels = 0
        for x in range(0, width, 5):  # Проверяем каждую 5-ю колонку
            r, g, b = pixels[x, y]
            if r < threshold and g < threshold and b < threshold:  # Темная линия
                line_pixels += 1
        # Если линия проходит через более 40% ширины - считаем её горизонтальной линией таблицы
        if line_pixels > (width // 5) * 0.4:
            horizontal_line_count += 1
    
    # Ищем вертикальные линии (проходят через большую часть высоты)
    for x in range(0, width, vertical_step):
        line_pixels = 0
        for y in range(0, height, 5):  # Проверяем каждую 5-ю строку
            r, g, b = pixels[x, y]
            if r < threshold and g < threshold and b < threshold:  # Темная линия
                line_pixels += 1
        # Если линия проходит через более 40% высоты - считаем её вертикальной линией таблицы
        if line_pixels > (height // 5) * 0.4:
            vertical_line_count += 1
    
    # Нормализуем по размеру изображения
    h_score = horizontal_line_count / (height / horizontal_step) if horizontal_step > 0 else 0
    v_score = vertical_line_count / (width / vertical_step) if vertical_step > 0 else 0
    
    return h_score, v_score


def _analyze_color_distribution(image: Image.Image) -> dict:
    """Анализирует распределение цветов на изображении.
    
    Диаграммы обычно имеют более разнообразную цветовую палитру (много разных цветов),
    а таблицы - в основном черный текст на белом фоне (мало цветов).
    """
    width, height = image.size
    pixels = image.load()
    
    # Собираем уникальные цвета (кроме почти белых)
    unique_colors = set()
    color_samples = 0
    
    # Выбираем каждый 20-й пиксель для ускорения
    step = max(1, min(20, width // 50))
    
    for x in range(0, width, step):
        for y in range(0, height, step):
            r, g, b = pixels[x, y]
            # Игнорируем почти белые пиксели (фон)
            if r < 240 or g < 240 or b < 240:
                # Квантуем цвета для группировки похожих оттенков
                r_q = (r // 16) * 16
                g_q = (g // 16) * 16
                b_q = (b // 16) * 16
                unique_colors.add((r_q, g_q, b_q))
                color_samples += 1
    
    unique_color_count = len(unique_colors)
    color_diversity = unique_color_count / max(color_samples, 1) if color_samples > 0 else 0
    
    return {
        'unique_colors': unique_color_count,
        'color_samples': color_samples,
        'diversity': color_diversity,
    }


def _is_chart_page(image: Image.Image, original_size: tuple[int, int]) -> bool:
    """Определяет, является ли страница диаграммой (а не таблицей).
    
    Комплексный анализ:
    1. Табличная структура (регулярные линии) - признак таблицы
    2. Цветовое разнообразие - диаграммы имеют больше цветов
    3. Распределение контента - таблицы равномерны, диаграммы сконцентрированы
    4. Размер после обрезки - диаграммы имеют больше белого пространства
    """
    logger = logging.getLogger(__name__)
    width, height = original_size
    
    # Конвертируем в RGB для анализа
    if image.mode != 'RGB':
        image = image.convert('RGB')
    
    # 1. Проверяем наличие табличной структуры (регулярные линии)
    h_score, v_score = _detect_table_structure(image)
    table_structure_score = (h_score + v_score) / 2.0
    
    # 2. Анализируем цветовое разнообразие
    color_analysis = _analyze_color_distribution(image)
    color_diversity = color_analysis['diversity']
    unique_colors = color_analysis['unique_colors']
    
    logger.info(f"Page {width}x{height}: table_score={table_structure_score:.3f}, colors={unique_colors}, diversity={color_diversity:.3f}")
    
    # КРИТЕРИЙ 1: Сильная табличная структура + мало цветов (<10) = точно таблица
    if table_structure_score > 0.12 and unique_colors < 10:
        logger.info(f"Rejected: strong table structure + low color diversity (table_score={table_structure_score:.3f}, colors={unique_colors})")
        return False
    
    # КРИТЕРИЙ 2: Умеренная табличная структура + очень мало цветов (<6) = таблица
    if table_structure_score > 0.08 and unique_colors < 6:
        logger.info(f"Rejected: moderate table structure + very low color diversity (table_score={table_structure_score:.3f}, colors={unique_colors})")
        return False
    
    # КРИТЕРИЙ 3: Мало цветов (<8) вообще - скорее всего таблица или текст, не диаграмма
    if unique_colors < 8:
        logger.info(f"Rejected: too few colors (colors={unique_colors}) - likely table or text, not a chart")
        return False
    
    # КРИТЕРИЙ 4: Если много уникальных цветов (>15) - вероятно диаграмма (независимо от структуры)
    if unique_colors > 15:
        logger.info(f"Accepted: high color diversity (colors={unique_colors}) - likely chart")
        return True
    
    # 3. Дополнительная проверка: анализируем распределение контента
    pixels = image.load()
    threshold = 240
    step = max(1, min(20, width // 50))
    
    # Подсчитываем контент по областям (верх, центр, низ)
    top_region = height // 3
    bottom_region = height * 2 // 3
    
    top_content = 0
    center_content = 0
    bottom_content = 0
    total_samples = 0
    
    for x in range(0, width, step):
        for y in range(0, height, step):
            total_samples += 1
            r, g, b = pixels[x, y]
            is_content = r < threshold or g < threshold or b < threshold
            if is_content:
                if y < top_region:
                    top_content += 1
                elif y < bottom_region:
                    center_content += 1
                else:
                    bottom_content += 1
    
    if total_samples == 0:
        return False
    
    content_distribution = [
        top_content / total_samples if total_samples > 0 else 0,
        center_content / total_samples if total_samples > 0 else 0,
        bottom_content / total_samples if total_samples > 0 else 0,
    ]
    distribution_variance = sum((d - sum(content_distribution)/3)**2 for d in content_distribution) / 3
    total_content_ratio = sum(content_distribution)
    
    logger.info(f"Content: top={content_distribution[0]:.2%}, center={content_distribution[1]:.2%}, bottom={content_distribution[2]:.2%}, variance={distribution_variance:.4f}, total={total_content_ratio:.2%}")
    
    # Очень равномерное распределение + много контента = таблица (даже если мало цветов или есть табличная структура)
    if distribution_variance < 0.001 and total_content_ratio > 0.35:
        logger.info(f"Rejected: uniform content distribution indicates table (variance={distribution_variance:.4f}, content={total_content_ratio:.2%}, colors={unique_colors})")
        return False
    
    # Равномерное распределение + очень много контента (>50%) = таблица
    if distribution_variance < 0.002 and total_content_ratio > 0.5:
        logger.info(f"Rejected: high content density with uniform distribution (variance={distribution_variance:.4f}, content={total_content_ratio:.2%}, colors={unique_colors})")
        return False
    
    # 4. Проверка через обрезку (диаграммы обычно имеют больше белого пространства)
    try:
        cropped = _crop_chart_area(image.copy())
        cropped_area = cropped.width * cropped.height
        original_area = width * height
        size_ratio = cropped_area / original_area if original_area > 0 else 1.0
        
        logger.info(f"After crop: size_ratio={size_ratio:.2%}, cropped={cropped.width}x{cropped.height}")
        
        # Если после обрезки осталось мало (много белого пространства) - вероятно диаграмма
        if size_ratio < 0.5:
            logger.info("Accepted: significant white space after crop (likely chart)")
            return True
    except Exception as e:
        logger.warning(f"Error during crop check: {e}")
    
    # 5. Итоговое решение на основе комбинации факторов
    # Если очень мало контента - пропускаем (пустая страница)
    if total_content_ratio < 0.05:
        logger.info("Rejected: too little content")
        return False
    
    # СТРОГИЕ КРИТЕРИИ для диаграмм: должны быть выполнены ОБА условия:
    # 1. Достаточно цветов (>=10)
    # 2. И (нет табличной структуры ИЛИ неравномерное распределение)
    if unique_colors >= 10:
        if table_structure_score < 0.15:
            logger.info(f"Accepted: good color diversity + weak table structure (colors={unique_colors}, table_score={table_structure_score:.3f})")
            return True
        if distribution_variance > 0.002:
            logger.info(f"Accepted: good color diversity + non-uniform distribution (colors={unique_colors}, variance={distribution_variance:.4f})")
            return True
    
    # Если очень много цветов (>18) - точно диаграмма
    if unique_colors > 18:
        logger.info(f"Accepted: very high color diversity (colors={unique_colors}) - definitely a chart")
        return True
    
    # В остальных случаях - отклоняем (вероятно таблица)
    # Особенно если мало цветов ИЛИ есть табличная структура ИЛИ равномерное распределение
    logger.info(f"Rejected: table-like characteristics (table_score={table_structure_score:.3f}, colors={unique_colors}, content={total_content_ratio:.2%}, variance={distribution_variance:.4f})")
    return False


def _render_excel_charts(pdf_bytes: bytes, file_id: str, pages: int, expected_charts_count: Optional[int] = None) -> tuple[Path, Path, int, Path]:
    """Рендерит страницы PDF как изображения, но сохраняет только диаграммы (фильтрует таблицы).
    
    Примечание: LibreOffice конвертирует весь Excel лист в PDF, включая таблицы и диаграммы.
    Мы пытаемся определить, какие страницы содержат диаграммы, а какие - таблицы,
    и сохраняем только диаграммы.
    """
    logger = logging.getLogger(__name__)
    charts_dir = _excel_charts_dir_path(file_id)
    charts_dir.mkdir(parents=True, exist_ok=True)
    preview_path: Optional[Path] = None
    width = height = None
    rendered_count = 0
    chart_index = 1  # Индекс для именования файлов диаграмм (пропускаем таблицы)
    
    if expected_charts_count:
        logger.info(f"Expected {expected_charts_count} chart(s) based on openpyxl detection - will stop after finding all charts")
    
    for page in range(1, pages + 1):
        # Если мы знаем количество диаграмм и уже нашли все - останавливаемся
        if expected_charts_count and rendered_count >= expected_charts_count:
            logger.info(f"Found all {expected_charts_count} expected chart(s), stopping page analysis")
            break
        
        chart_bytes = _render_pdf_page(pdf_bytes, page, scale=1.2)
        if not chart_bytes:
            logger.warning(f"Failed to render Excel page {page} for {file_id}")
            continue
        
        image = Image.open(BytesIO(chart_bytes))
        original_size = image.size
        
        # Определяем, является ли страница диаграммой
        # Если знаем количество диаграмм, можем быть более строгими к таблицам
        is_chart = _is_chart_page(image, original_size)
        
        if not is_chart:
            logger.info(f"Page {page}/{pages} for {file_id} - REJECTED (identified as table)")
            continue  # Пропускаем таблицы, сохраняем только диаграммы
        
        logger.info(f"Page {page}/{pages} for {file_id} - ACCEPTED as chart #{chart_index} ({rendered_count + 1}/{expected_charts_count if expected_charts_count else '?'})")
        
        # Обрезаем изображение, оставляя только контент (диаграмму)
        try:
            cropped_image = _crop_chart_area(image)
        except Exception as e:
            logger.warning(f"Failed to crop chart image for page {page}: {e}")
            cropped_image = image  # Используем оригинал, если обрезка не удалась
        
        if EXCEL_CHARTS_TARGET_WIDTH and cropped_image.width > EXCEL_CHARTS_TARGET_WIDTH:
            cropped_image.thumbnail((EXCEL_CHARTS_TARGET_WIDTH, EXCEL_CHARTS_TARGET_WIDTH * 2), Image.Resampling.LANCZOS)
        buffer = BytesIO()
        cropped_image.save(buffer, format="WEBP", quality=85)
        chart_bytes = buffer.getvalue()
        chart_path = charts_dir / f"{chart_index}.webp"
        _write_file(chart_path, chart_bytes)
        chart_index += 1
        rendered_count += 1
        if preview_path is None:
            preview_path = chart_path
            width, height = cropped_image.size
    
    if preview_path is None:
        logger.info(f"No charts found in Excel file {file_id} (only tables or empty pages)")
        return None  # Возвращаем None, если не найдено ни одной диаграммы
    
    meta = {
        "count": rendered_count,
        "format": "webp",
        "width": width,
        "height": height,
    }
    json_path = _excel_charts_json_path(file_id)
    json_path.write_text(json.dumps(meta), encoding="utf-8")
    logger.info("Generated %s chart pages (filtered out tables) for %s", rendered_count, file_id)
    return preview_path, json_path, rendered_count, charts_dir


def _build_block(asset: FileAsset) -> dict:
    if asset.kind == "image":
        block = ImageBlock(
            type="image",
            id=generate_uuid(),
            data=ImageData(
                src=_file_url(asset.id, "preview") if asset.path_preview else _file_url(asset.id, "original"),
                full=_file_url(asset.id, "original"),
                alt=asset.filename,
                w=asset.width,
                h=asset.height,
            ),
        )
        return dump_block(block)

    if asset.kind == "pdf":
        # Извлекаем имя файла без расширения для title
        filename_without_ext = Path(asset.filename).stem if asset.filename else None
        doc_data = DocData(
            kind="pdf",
            src=_file_url(asset.id, "original"),
            title=filename_without_ext or "PDF-документ",
            preview=_file_url(asset.id, "preview") if asset.path_preview else None,
            meta=DocMeta(pages=asset.pages, slides=None, size=asset.size),
            view="cover",  # OVC: pdf - по умолчанию режим обложки
        )
        block = DocBlock(type="doc", id=generate_uuid(), data=doc_data)
        return dump_block(block)

    if asset.kind in {"docx", "rtf"}:
        filename_without_ext = Path(asset.filename).stem if asset.filename else None
        doc_data = DocData(
            kind="docx" if asset.kind == "docx" else "rtf",
            src=_file_url(asset.id, "original"),
            title=filename_without_ext or asset.filename,
            preview=_file_url(asset.id, "preview") if asset.path_preview else None,
            meta=DocMeta(pages=asset.pages, slides=None, size=asset.size, words=asset.words),
            view="cover",
        )
        block = DocBlock(type="doc", id=generate_uuid(), data=doc_data)
        return dump_block(block)

    if asset.kind == "audio":
        audio_data = AudioData(
            src=_file_url(asset.id, "stream"),
            mime=asset.mime,
            duration=asset.duration,
            waveform=_file_url(asset.id, "waveform") if asset.path_waveform else None,
            transcript=None,
            view="mini",
        )
        block = AudioBlock(type="audio", id=generate_uuid(), data=audio_data)
        return dump_block(block)
    if asset.kind == "video":
        video_data = VideoData(
            src=_video_source_url(asset.id),
            poster=_video_poster_url(asset.id) if asset.path_video_poster else None,
            duration_sec=asset.video_duration,
            width=asset.video_width,
            height=asset.video_height,
            mime=asset.video_mime or asset.mime,
            caption="",
            view="cover",
        )
        block = VideoBlock(type="video", id=generate_uuid(), data=video_data)
        return dump_block(block)

    if asset.kind == "pptx":
        # OVC: pptx - если слайды не были сгенерированы (LibreOffice не установлен), все равно создаем блок
        slides_data = SlidesData(
            kind="pptx",
            src=_file_url(asset.id, "original"),
            slides=_slides_json_url(asset.id) if asset.path_slides_json else None,
            preview=_slide_image_url(asset.id, 1) if asset.path_preview else None,
            count=asset.slides_count,
            view="cover",
        )
        block = SlidesBlock(type="slides", id=generate_uuid(), data=slides_data)
        return dump_block(block)

    if asset.kind in {"xlsx", "xls", "csv"}:
        # OVC: excel - диаграммы отключены, фокус на предпросмотре таблиц
        table_data = TableData(
            kind=asset.kind,
            src=_file_url(asset.id, "original"),
            summary=_excel_summary_url(asset.id),
            view="cover",
            active_sheet=asset.excel_default_sheet,
            charts=None,  # Диаграммы отключены
        )
        block = TableBlock(type="table", id=generate_uuid(), data=table_data)
        return dump_block(block)
    if asset.kind == "markdown":
        md_data = MarkdownData(
            src=_markdown_raw_url(asset.id),
            previewUrl=_markdown_preview_url(asset.id),
            filename=asset.filename,
            sizeBytes=asset.size,
            lineCount=asset.markdown_line_count,
            view="inline",
        )
        block = MarkdownBlock(type="markdown", id=generate_uuid(), data=md_data)
        return dump_block(block)
    if asset.kind == "code":
        code_data = CodeData(
            src=_code_raw_url(asset.id),
            previewUrl=_code_preview_url(asset.id),
            filename=asset.filename,
            language=asset.code_language or "plaintext",
            sizeBytes=asset.size,
            lineCount=asset.code_line_count,
            view="inline",
        )
        block = CodeBlock(type="code", id=generate_uuid(), data=code_data)
        return dump_block(block)

    raise HTTPException(status_code=500, detail=f"Unknown asset kind: {asset.kind}")


@dataclass
class StoredAsset:
    asset: FileAsset
    block: dict


async def save_upload(
    session: Session,
    upload: UploadFile,
    note_id: Optional[str],
    user_id: str,
    upload_op_id: Optional[str] = None,
) -> StoredAsset:
    logger = logging.getLogger(__name__)  # OVC: docx - определяем logger в начале функции
    meta = _classify_file(upload)
    data = await upload.read()
    if not data:
        raise HTTPException(status_code=400, detail="Empty file")
    if len(data) > meta.max_bytes:
        raise HTTPException(status_code=413, detail="File too large for this prototype")

    file_id = generate_uuid()
    original_name = upload.filename or f"{file_id}{meta.extension}"
    if meta.kind == "video":
        original_path = _video_original_path(file_id, meta.extension)
    elif meta.kind == "code":
        original_path = _code_original_path(file_id, meta.extension)
    elif meta.kind == "markdown":
        original_path = _markdown_raw_path(file_id, meta.extension)
    else:
        original_path = ORIGINAL_DIR / f"{file_id}{meta.extension}"
    _write_file(original_path, data)

    preview_path: Optional[Path] = None
    doc_html_path: Optional[Path] = None
    waveform_path: Optional[Path] = None
    slides_json_path: Optional[Path] = None
    slides_dir_path: Optional[Path] = None
    summary_path: Optional[Path] = None
    default_sheet: Optional[str] = None
    width = None
    height = None
    pages = None
    words = None
    duration = None
    slides_count = None
    video_poster_path: Optional[Path] = None
    video_duration = None
    video_width = None
    video_height = None
    video_mime = None
    code_language = None
    code_line_count = None
    markdown_line_count = None
    # OVC: excel - инициализируем переменные для диаграмм (нужны для всех типов файлов)
    charts_json_path: Optional[str] = None
    charts_dir_path: Optional[str] = None
    charts_sheets_json_path: Optional[str] = None
    charts_count: Optional[int] = None

    if meta.kind == "image":
        try:
            preview_bytes, width, height = _generate_image_preview(data)
            preview_path = PREVIEW_DIR / f"{file_id}.webp"
            _write_file(preview_path, preview_bytes)
        except Exception as exc:
            logger.warning(f"Image preview generation failed for {file_id}: {exc}")
            preview_path = None
    elif meta.kind == "pdf":
        pages = _pdf_page_count(data)
        preview_bytes = _generate_pdf_preview(data, file_id, pages)  # OVC: pdf - передаем data для рендеринга
        preview_path = PREVIEW_DIR / f"{file_id}.webp"
        _write_file(preview_path, preview_bytes)
    elif meta.kind in {"docx", "rtf"}:
        try:
            if meta.kind == "docx":
                html_fragment, words = _convert_docx_to_html(data)
            else:
                html_fragment, words = _convert_rtf_to_html(data)
            sanitized = _sanitize_html_fragment(html_fragment)
            doc_html_path = _write_doc_html(file_id, sanitized)
            
            # OVC: docx - не генерируем превью для Word файлов, используем только inline просмотр
            # Превью не создается - будет показан бейдж с типом файла
            preview_path = None
            logger.info(f"Processed {meta.kind} file {file_id}, HTML length: {len(html_fragment)}, words: {words}")
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"Error processing {meta.kind} file: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail=f"Error processing {meta.kind} file: {str(e)}")
    elif meta.kind == "pptx":
        # OVC: pptx - пытаемся конвертировать, но не падаем если LibreOffice не установлен
        try:
            preview_path, slides_json_path, slides_count, slides_dir_path = _convert_pptx_to_slides(original_path, file_id)
            doc_html_path = None
        except HTTPException as e:
            # Если LibreOffice не установлен или конвертация не удалась, загружаем файл без preview
            logger.warning(f"PPTX conversion failed for {file_id}: {e.detail}. File will be saved without preview.")
            preview_path = None
            slides_json_path = None
            slides_dir_path = None
            slides_count = None
            doc_html_path = None
    elif meta.kind in {"xlsx", "xls", "csv"}:
        summary_path, default_sheet = _prepare_excel_summary(meta.kind, original_path, file_id)
        doc_html_path = None
        # OVC: excel - диаграммы отключены, фокус на предпросмотре таблиц
    elif meta.kind == "audio":
        duration, waveform_values = _extract_audio_metadata(data, meta.mime)
        waveform_path = _write_waveform(file_id, waveform_values)
    elif meta.kind == "video":
        video_mime = meta.mime
        meta_info = extract_video_meta(original_path)
        video_duration = meta_info.get("duration")
        video_width = meta_info.get("width")
        video_height = meta_info.get("height")
        poster_candidate = _video_poster_path(file_id)
        if render_video_poster(original_path, poster_candidate, duration=video_duration):
            video_poster_path = poster_candidate
    elif meta.kind == "code":
        code_language = _detect_code_language(meta.extension)
        try:
            code_line_count = count_file_lines(original_path)
        except Exception:
            code_line_count = None
    elif meta.kind == "markdown":
        try:
            markdown_line_count = count_file_lines(original_path)
        except Exception:
            markdown_line_count = None

    asset = FileAsset(
        id=file_id,
        note_id=note_id,
        user_id=user_id,
        kind=meta.kind,
        mime=meta.mime,
        filename=original_name,
        size=len(data),
        path_original=str(original_path),
        path_preview=str(preview_path) if preview_path else None,
        path_doc_html=str(doc_html_path) if doc_html_path else None,
        path_slides_json=str(slides_json_path) if slides_json_path else None,
        path_slides_dir=str(slides_dir_path) if slides_dir_path else None,
        path_waveform=str(waveform_path) if waveform_path else None,
        path_excel_summary=str(summary_path) if summary_path else None,
        path_video_original=str(original_path) if meta.kind == "video" else None,
        path_video_poster=str(video_poster_path) if video_poster_path else None,
        path_code_original=str(original_path) if meta.kind == "code" else None,
        path_markdown_raw=str(original_path) if meta.kind == "markdown" else None,
        path_excel_charts_json=charts_json_path,
        path_excel_charts_dir=charts_dir_path,
        path_excel_chart_sheets_json=charts_sheets_json_path,
        excel_default_sheet=default_sheet,
        hash_sha256=_hash_bytes(data),
        upload_op_id=upload_op_id,
        width=width,
        height=height,
        pages=pages,
        duration=duration,
        words=words,
        slides_count=slides_count,
        video_duration=video_duration,
        video_width=video_width,
        video_height=video_height,
        video_mime=video_mime,
        code_language=code_language,
        code_line_count=code_line_count,
        markdown_line_count=markdown_line_count,
    )
    session.add(asset)
    session.flush()

    block = _build_block(asset)
    return StoredAsset(asset=asset, block=block)
